from collections import defaultdict
from typing import Dict, Optional, Tuple, List

import pandas as pd


# =========================================================
# 读数与标准化
# =========================================================

def load_data(
    demand_path: str,
    hist_path: str,
    aoi_cap_path: str,
    cut_cap_path: str,
    changeover_path: str,
    hist_recipe_path: Optional[str] = None,
    irregular_path: Optional[str] = None,
    year: int = 2026,
):
    demand_raw = pd.read_csv(demand_path)
    hist_raw = pd.read_csv(hist_path)
    aoi_raw = pd.read_csv(aoi_cap_path)
    cut_raw = pd.read_csv(cut_cap_path)
    changeover_raw = pd.read_csv(changeover_path)
    hist_recipe_raw = pd.read_csv(hist_recipe_path) if hist_recipe_path else pd.DataFrame(columns=["PROCESS", "LINE", "MODEL_NO"])
    irregular_raw = pd.read_csv(irregular_path) if irregular_path else pd.DataFrame()

    # ---------- demand ----------
    demand = demand_raw.dropna(subset=["MFG_DAY", "MODEL_NO", "PART_NO", "PPC_DPS"]).copy()
    demand = demand[demand["PPC_DPS"].fillna(0) > 0].copy()
    demand["date"] = pd.to_datetime(demand["MFG_DAY"], errors="coerce")
    if demand["date"].isna().all():
        demand["date"] = pd.to_datetime(f"{year}-" + demand["MFG_DAY"].astype(str), format="%Y-%m-%d", errors="coerce")
    demand = demand[demand["date"].notna()].copy()
    demand["MODEL_NO"] = demand["MODEL_NO"].astype(str).str.strip()
    demand["PART_NO"] = demand["PART_NO"].astype(str).str.strip()
    demand["key"] = demand["MODEL_NO"] + "|" + demand["PART_NO"]
    demand = demand[["date", "MODEL_NO", "PART_NO", "key", "PPC_DPS"]].rename(columns={"PPC_DPS": "qty"})

    # ---------- history ----------
    hist = hist_raw.copy()
    keep = [c for c in ["PROCESS", "LINE", "MODEL_NO", "PART_NO"] if c in hist.columns]
    hist = hist[keep].copy()
    hist["PROCESS"] = hist["PROCESS"].astype(str).str.strip().str.upper()
    hist["LINE"] = hist["LINE"].astype(str).str.strip()

    def _norm_optional_text(x):
        if pd.isna(x) or x is None:
            return None
        s = str(x).strip()
        if s == "" or s.lower() in {"nan", "none", "null"}:
            return None
        return s

    hist["MODEL_NO"] = hist["MODEL_NO"].apply(_norm_optional_text)
    hist["PART_NO"] = hist["PART_NO"].apply(_norm_optional_text)

    def make_key(m, p):
        if pd.isna(m) or m is None or pd.isna(p) or p is None:
            return None
        return f"{str(m).strip()}|{str(p).strip()}"

    hist["key"] = hist.apply(lambda r: make_key(r["MODEL_NO"], r["PART_NO"]), axis=1)

    # ---------- CUT capability ----------
    cut_id_cols = [c for c in ["MODEL_NO", "PART_NO", "RATIO"] if c in cut_raw.columns]
    cut_cap = cut_raw.melt(id_vars=cut_id_cols, var_name="line", value_name="cap_shift")
    cut_cap = cut_cap[cut_cap["cap_shift"].fillna(0) > 0].copy()
    cut_cap["MODEL_NO"] = cut_cap["MODEL_NO"].astype(str).str.strip()
    cut_cap["PART_NO"] = cut_cap["PART_NO"].astype(str).str.strip()
    cut_cap["line"] = cut_cap["line"].astype(str).str.strip()
    cut_cap["key"] = cut_cap["MODEL_NO"] + "|" + cut_cap["PART_NO"]
    cut_cap["RATIO"] = cut_cap["RATIO"].astype(str).str.strip()

    def _parse_ratio_pct(x):
        s = str(x).strip().replace(" ", "")
        if not s:
            return None
        if s.endswith("%"):
            s = s[:-1]
        try:
            return float(s)
        except Exception:
            return None

    cut_cap["ratio_pct"] = cut_cap["RATIO"].apply(_parse_ratio_pct)
    grp = cut_cap.groupby(["MODEL_NO", "PART_NO", "key", "line"])["ratio_pct"].transform(lambda s: s.notna().sum())
    cut_cap["ratio_rows"] = grp.fillna(0).astype(int)
    cut_cap["ratio_is_100"] = (cut_cap["ratio_rows"] <= 1) & cut_cap["ratio_pct"].fillna(100.0).round(6).eq(100.0)
    cut_cap["ratio_frac"] = cut_cap["ratio_pct"].fillna(100.0) / 100.0
    cut_cap["cap_day"] = 2.0 * cut_cap["cap_shift"]

    # ---------- AOI capability ----------
    aoi_id_cols = [c for c in ["MODEL_NO", "MAX_LINE"] if c in aoi_raw.columns]
    aoi_cap = aoi_raw.melt(id_vars=aoi_id_cols, var_name="line", value_name="cap_shift")
    aoi_cap = aoi_cap[aoi_cap["cap_shift"].fillna(0) > 0].copy()
    aoi_cap["MODEL_NO"] = aoi_cap["MODEL_NO"].astype(str).str.strip()
    aoi_cap["line"] = aoi_cap["line"].astype(str).str.strip()
    aoi_cap["cap_day"] = 2.0 * aoi_cap["cap_shift"]

    # ---------- changeover ----------
    changeover = changeover_raw.copy()
    changeover["PROCESS"] = changeover["PROCESS"].astype(str).str.strip().str.upper()
    changeover["LINE"] = changeover["LINE"].astype(str).str.strip()
    if "IF_REGULAR" in changeover.columns:
        changeover["IF_REGULAR"] = changeover["IF_REGULAR"].astype(str).str.strip().str.upper()
    for col in ["WITH_RE", "WO_RE", "WITH_RATIO", "PART_CHANGE"]:
        if col in changeover.columns:
            changeover[col] = pd.to_numeric(changeover[col], errors="coerce")

    # ---------- hist recipe ----------
    hist_recipe = hist_recipe_raw.copy()
    keep_recipe = [c for c in ["PROCESS", "LINE", "MODEL_NO"] if c in hist_recipe.columns]
    hist_recipe = hist_recipe[keep_recipe].copy() if keep_recipe else pd.DataFrame(columns=["PROCESS", "LINE", "MODEL_NO"])
    if not hist_recipe.empty:
        hist_recipe["PROCESS"] = hist_recipe["PROCESS"].astype(str).str.strip().str.upper()
        hist_recipe["LINE"] = hist_recipe["LINE"].astype(str).str.strip()
        hist_recipe["MODEL_NO"] = hist_recipe["MODEL_NO"].astype(str).str.strip()

    # ---------- irregular ----------
    irregular_models = set()
    if irregular_raw is not None and not irregular_raw.empty:
        for c in irregular_raw.columns:
            v = str(c).strip()
            if v and v.lower() != "nan":
                irregular_models.add(v)
        first_col = irregular_raw.columns[0]
        irregular_series = irregular_raw[first_col].dropna().astype(str).str.strip()
        irregular_models.update([x for x in irregular_series.tolist() if x and x.lower() != "nan"])

    return {
        "demand": demand,
        "hist": hist,
        "cut_cap": cut_cap,
        "aoi_cap": aoi_cap,
        "changeover": changeover,
        "hist_recipe": hist_recipe,
        "irregular_models": irregular_models,
    }



def step0_checks(data: Dict) -> Tuple[pd.DataFrame, pd.DataFrame]:
    demand = data["demand"]
    hist = data["hist"]
    cut_cap = data["cut_cap"]
    aoi_cap = data["aoi_cap"]

    summary = pd.DataFrame(
        [{
            "demand_rows_after_clean": len(demand),
            "days_with_demand": demand["date"].nunique(),
            "unique_products": demand["key"].nunique(),
            "unique_models": demand["MODEL_NO"].nunique(),
            "hist_rows": len(hist),
            "hist_cut_rows": int((hist["PROCESS"] == "CUT").sum()),
            "hist_aoi_rows": int((hist["PROCESS"] == "AOI").sum()),
        }]
    )

    warnings = []
    next_date = pd.Timestamp("2026-02-04")

    for _, r in hist.iterrows():
        process = r["PROCESS"]
        line = r["LINE"]
        model = r["MODEL_NO"]
        part = r["PART_NO"]
        key = r["key"]

        if model is None:
            continue

        if process == "CUT":
            if key is not None:
                x = demand[(demand["date"] == next_date) & (demand["key"] == key)]
                if len(x) > 0:
                    cap_ok = ((cut_cap["key"] == key) & (cut_cap["line"] == line)).any()
                    if not cap_ok:
                        warnings.append({
                            "date": str(next_date.date()),
                            "process": "CUT",
                            "line": line,
                            "model": model,
                            "part": part,
                            "msg": "历史-能力不一致告警",
                        })

        elif process == "AOI":
            x = demand[(demand["date"] == next_date) & (demand["MODEL_NO"] == model)]
            if len(x) > 0:
                cap_ok = ((aoi_cap["MODEL_NO"] == model) & (aoi_cap["line"] == line)).any()
                if not cap_ok:
                    warnings.append({
                        "date": str(next_date.date()),
                        "process": "AOI",
                        "line": line,
                        "model": model,
                        "part": None,
                        "msg": "历史-能力不一致告警",
                    })

    return summary, pd.DataFrame(warnings)


# =========================================================
# 通用函数
# =========================================================

def _has_recipe_run_before(
    process: str,
    line: Optional[str],
    model: Optional[str],
    hist_recipe_df: Optional[pd.DataFrame],
) -> bool:
    if hist_recipe_df is None or hist_recipe_df.empty or line is None or model is None:
        return False
    process = str(process).strip().upper()
    line = str(line).strip()
    model = str(model).strip()
    m = (
        (hist_recipe_df["PROCESS"] == process)
        & (hist_recipe_df["LINE"] == line)
        & (hist_recipe_df["MODEL_NO"] == model)
    )
    return bool(m.any())


def _get_changeover_row(
    process: str,
    line: Optional[str],
    changeover_df: pd.DataFrame,
    if_regular: Optional[bool] = None,
) -> pd.Series:
    process = str(process).strip().upper()
    line = None if line is None else str(line).strip()

    df = changeover_df[changeover_df["PROCESS"] == process].copy()
    if line is not None and "LINE" in df.columns:
        df = df[df["LINE"] == line].copy()

    if process == "CUT" and "IF_REGULAR" in df.columns and if_regular is not None:
        regular_tag = "YES" if if_regular else "NO"
        df2 = df[df["IF_REGULAR"] == regular_tag].copy()
        if not df2.empty:
            df = df2

    if df.empty:
        raise ValueError(f"changeover row not found for process={process}, line={line}, if_regular={if_regular}")
    return df.iloc[0]


def get_changeover_h(
    process: str,
    from_model: Optional[str],
    from_part: Optional[str],
    to_model: str,
    to_part: Optional[str],
    changeover_df: pd.DataFrame,
    line: Optional[str] = None,
    hist_recipe_df: Optional[pd.DataFrame] = None,
    irregular_models: Optional[set] = None,
    ratio_is_100: Optional[bool] = None,
) -> float:
    process = str(process).upper()

    if process == "CUT":
        if from_model == to_model and from_part == to_part:
            return 0.0

        irregular_models = irregular_models or set()
        is_regular = to_model not in irregular_models
        row = _get_changeover_row("CUT", line=line, changeover_df=changeover_df, if_regular=is_regular)

        ran_before = _has_recipe_run_before("CUT", line=line, model=to_model, hist_recipe_df=hist_recipe_df)
        col = "WITH_RE" if ran_before else "WO_RE"
        base = row.get(col)
        if pd.notna(base):
            return float(base)
        return 2.0

    if process == "AOI":
        row = _get_changeover_row("AOI", line=line, changeover_df=changeover_df)

        if from_model == to_model:
            if from_part == to_part:
                return 0.0
            part_change = row.get("PART_CHANGE")
            return float(part_change) if pd.notna(part_change) else 0.0

        ran_before = _has_recipe_run_before("AOI", line=line, model=to_model, hist_recipe_df=hist_recipe_df)
        col = "WITH_RE" if ran_before else "WO_RE"
        base = row.get(col)
        return float(base) if pd.notna(base) else 8.0

    raise ValueError(process)



def get_cut_ratio_changeover_h(
    to_model: str,
    line: Optional[str],
    changeover_df: pd.DataFrame,
    irregular_models: Optional[set] = None,
) -> float:
    irregular_models = irregular_models or set()
    is_regular = to_model not in irregular_models
    row = _get_changeover_row("CUT", line=line, changeover_df=changeover_df, if_regular=is_regular)
    v = row.get("WITH_RATIO")
    return float(v) if pd.notna(v) else 0.0


def get_cut_segments(cap_info: pd.DataFrame) -> List[Dict]:
    if cap_info is None or cap_info.empty:
        return []
    seg = cap_info.copy()
    if "ratio_frac" not in seg.columns:
        seg["ratio_frac"] = 1.0
    if "ratio_pct" not in seg.columns:
        seg["ratio_pct"] = seg["ratio_frac"] * 100.0
    seg = seg.sort_values(["ratio_frac", "cap_shift"], ascending=[False, False]).reset_index(drop=True)
    rows = []
    total_frac = float(seg["ratio_frac"].sum()) if not seg.empty else 0.0
    if total_frac <= 0:
        cap_shift = float(seg["cap_shift"].max()) if not seg.empty else 0.0
        rows.append({"ratio_frac": 1.0, "ratio_pct": 100.0, "cap_shift": cap_shift, "ratio_is_100": True})
        return rows
    for _, r in seg.iterrows():
        rows.append({
            "ratio_frac": float(r.get("ratio_frac", 0.0)) / total_frac,
            "ratio_pct": float(r.get("ratio_pct", 0.0)),
            "cap_shift": float(r.get("cap_shift", 0.0)),
            "ratio_is_100": bool(r.get("ratio_is_100", False)),
        })
    return rows


def get_first_changeover_h_for_line(
    process: str,
    line: str,
    target_model: str,
    target_part: Optional[str],
    line_state: Dict[str, Dict[str, Dict]],
    changeover_df: pd.DataFrame,
    hist_recipe_df: Optional[pd.DataFrame] = None,
    irregular_models: Optional[set] = None,
) -> float:
    prev = line_state.get(process, {}).get(line, {})
    return get_changeover_h(
        process=process,
        from_model=prev.get("model"),
        from_part=prev.get("part"),
        to_model=target_model,
        to_part=target_part,
        changeover_df=changeover_df,
        line=line,
        hist_recipe_df=hist_recipe_df,
        irregular_models=irregular_models,
    )


def get_effective_cap_day_after_first_changeover(
    cap_shift: float,
    first_changeover_h: float,
) -> float:
    day_qty, night_qty = plan_shift_qty_with_changeover_shift(cap_shift, first_changeover_h, "day")
    return float(day_qty + night_qty)


def get_segment_capacity_contribution(
    effective_cap_day_today: float,
    raw_cap_day: float,
    seg_len: int,
) -> float:
    seg_len = max(int(seg_len), 1)
    return float(effective_cap_day_today + raw_cap_day * max(seg_len - 1, 0))


def get_global_cut_ratio_signature(cut_cap: pd.DataFrame, key: str) -> Tuple[float, ...]:
    if cut_cap is None or cut_cap.empty:
        return tuple()
    sub = cut_cap[cut_cap["key"] == key].copy()
    if sub.empty or "ratio_pct" not in sub.columns:
        return tuple()
    vals = []
    for x in sub["ratio_pct"].dropna().tolist():
        try:
            vals.append(round(float(x), 6))
        except Exception:
            pass
    vals = sorted(set(vals))
    if not vals:
        return tuple((100.0,))
    return tuple(vals)


def line_supports_full_cut_ratio_structure(cut_cap: pd.DataFrame, key: str, line: str) -> bool:
    global_sig = get_global_cut_ratio_signature(cut_cap, key)
    if len(global_sig) <= 1:
        return True
    sub = cut_cap[(cut_cap["key"] == key) & (cut_cap["line"] == line)].copy()
    if sub.empty:
        return False
    line_vals = []
    for x in sub["ratio_pct"].dropna().tolist():
        try:
            line_vals.append(round(float(x), 6))
        except Exception:
            pass
    line_sig = tuple(sorted(set(line_vals)))
    return line_sig == global_sig


def filter_cut_lines_supporting_full_ratio(cut_cap: pd.DataFrame, key: str, lines: List[str]) -> List[str]:
    return [ln for ln in lines if line_supports_full_cut_ratio_structure(cut_cap, key, ln)]


def append_hist_recipe_if_new(hist_recipe_df: pd.DataFrame, process: str, line: str, model: Optional[str]) -> pd.DataFrame:
    if hist_recipe_df is None or model is None or str(model).strip() == "":
        return hist_recipe_df
    if hist_recipe_df.empty:
        return pd.DataFrame([{"PROCESS": process, "LINE": line, "MODEL_NO": model}])
    exists = (
        hist_recipe_df["PROCESS"].astype(str).str.upper().eq(str(process).upper()) &
        hist_recipe_df["LINE"].astype(str).eq(str(line)) &
        hist_recipe_df["MODEL_NO"].astype(str).eq(str(model))
    ).any()
    if not exists:
        hist_recipe_df = pd.concat(
            [hist_recipe_df, pd.DataFrame([{"PROCESS": process, "LINE": line, "MODEL_NO": model}])],
            ignore_index=True,
        )
    return hist_recipe_df


def get_production_dates(demand: pd.DataFrame) -> List[pd.Timestamp]:
    # 生产日序列：默认按排产窗口内“非周日”日期生成，而不是只取有需求的自然日
    # 这样可以支持：周六-周一连续；以及 CUT backlog 在无新需求的工作日继续生产
    cal_start = demand.attrs.get("calendar_start")
    cal_end = demand.attrs.get("calendar_end")
    if cal_start is not None and cal_end is not None:
        rng = pd.date_range(start=pd.Timestamp(cal_start).normalize(), end=pd.Timestamp(cal_end).normalize(), freq="D")
        dates = [pd.Timestamp(d) for d in rng if pd.Timestamp(d).weekday() != 6]
        return dates

    pos_dates = pd.to_datetime(demand.loc[demand["qty"] > 0, "date"]).dt.normalize()
    if len(pos_dates) == 0:
        return []
    rng = pd.date_range(start=pos_dates.min(), end=pos_dates.max(), freq="D")
    dates = [pd.Timestamp(d) for d in rng if pd.Timestamp(d).weekday() != 6]
    return dates


def get_future_production_dates(
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    max_slots: int = 3,
) -> List[pd.Timestamp]:
    production_dates = get_production_dates(demand)
    current_date = pd.Timestamp(current_date).normalize()
    future = [d for d in production_dates if d >= current_date]
    return future[:max_slots]


def get_next_production_date(demand: pd.DataFrame, current_date: pd.Timestamp) -> Optional[pd.Timestamp]:
    production_dates = get_production_dates(demand)
    current_date = pd.Timestamp(current_date).normalize()
    future = [d for d in production_dates if d > current_date]
    return future[0] if future else None


def _positive_qty_on_product(demand: pd.DataFrame, d: pd.Timestamp, key: str) -> float:
    x = demand[(demand["date"] == d) & (demand["key"] == key)]
    return float(x["qty"].sum()) if len(x) > 0 else 0.0


def _positive_qty_on_model(demand: pd.DataFrame, d: pd.Timestamp, model: str) -> float:
    x = demand[(demand["date"] == d) & (demand["MODEL_NO"] == model)]
    return float(x["qty"].sum()) if len(x) > 0 else 0.0


def get_sticky_segment_dates_product(
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    key: str,
    max_days: int = 3,
    gap_tolerance: int = 1,
) -> List[pd.Timestamp]:
    slots = get_future_production_dates(demand, current_date, max_slots=max_days)
    active_dates = []
    misses = 0
    seen_positive = False
    for d in slots:
        qty = _positive_qty_on_product(demand, d, key)
        if qty > 0:
            active_dates.append(d)
            seen_positive = True
            misses = 0
        else:
            if not seen_positive:
                break
            misses += 1
            if misses > gap_tolerance:
                break
    return active_dates


def get_sticky_segment_dates_model(
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    model: str,
    max_days: int = 3,
    gap_tolerance: int = 1,
) -> List[pd.Timestamp]:
    slots = get_future_production_dates(demand, current_date, max_slots=max_days)
    active_dates = []
    misses = 0
    seen_positive = False
    for d in slots:
        qty = _positive_qty_on_model(demand, d, model)
        if qty > 0:
            active_dates.append(d)
            seen_positive = True
            misses = 0
        else:
            if not seen_positive:
                break
            misses += 1
            if misses > gap_tolerance:
                break
    return active_dates


def next_day_has_same_product(demand: pd.DataFrame, current_date: pd.Timestamp, key: str) -> bool:
    dates = get_sticky_segment_dates_product(demand, current_date, key, max_days=3, gap_tolerance=1)
    return len(dates) >= 2


def next_day_has_same_model(demand: pd.DataFrame, current_date: pd.Timestamp, model: str) -> bool:
    dates = get_sticky_segment_dates_model(demand, current_date, model, max_days=3, gap_tolerance=1)
    return len(dates) >= 2


def consecutive_segment_len_product(demand: pd.DataFrame, current_date: pd.Timestamp, key: str, max_days: int = 3) -> int:
    return len(get_sticky_segment_dates_product(demand, current_date, key, max_days=max_days, gap_tolerance=1))


def consecutive_segment_len_model(demand: pd.DataFrame, current_date: pd.Timestamp, model: str, max_days: int = 3) -> int:
    return len(get_sticky_segment_dates_model(demand, current_date, model, max_days=max_days, gap_tolerance=1))


def segment_total_product(demand: pd.DataFrame, current_date: pd.Timestamp, key: str, max_days: int = 3) -> float:
    dates = get_sticky_segment_dates_product(demand, current_date, key, max_days=max_days, gap_tolerance=1)
    return float(sum(_positive_qty_on_product(demand, d, key) for d in dates))


def segment_total_model(demand: pd.DataFrame, current_date: pd.Timestamp, model: str, max_days: int = 3) -> float:
    dates = get_sticky_segment_dates_model(demand, current_date, model, max_days=max_days, gap_tolerance=1)
    return float(sum(_positive_qty_on_model(demand, d, model) for d in dates))


def get_segment_deadline_product(demand: pd.DataFrame, current_date: pd.Timestamp, key: str, max_days: int = 3) -> pd.Timestamp:
    dates = get_sticky_segment_dates_product(demand, current_date, key, max_days=max_days, gap_tolerance=1)
    return dates[-1] if dates else pd.Timestamp(current_date).normalize()


def get_segment_deadline_model(demand: pd.DataFrame, current_date: pd.Timestamp, model: str, max_days: int = 3) -> pd.Timestamp:
    dates = get_sticky_segment_dates_model(demand, current_date, model, max_days=max_days, gap_tolerance=1)
    return dates[-1] if dates else pd.Timestamp(current_date).normalize()


def build_model_part_reference(demand: pd.DataFrame) -> Dict[str, str]:
    ref = {}
    tmp = demand.groupby("MODEL_NO")["PART_NO"].agg(
        lambda s: sorted(set([x for x in s.dropna().astype(str)]))
    )
    for model, parts in tmp.items():
        if len(parts) == 1:
            ref[model] = parts[0]
        elif len(parts) == 0:
            ref[model] = None
        else:
            ref[model] = "MULTI_PART"
    return ref


def get_day_model_part_map(
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    model_part_ref: Dict[str, str]
) -> Dict[str, str]:
    day = demand[demand["date"] == current_date].copy()
    if day.empty:
        return {}

    out = {}
    tmp = day.groupby("MODEL_NO")["PART_NO"].agg(
        lambda s: sorted(set([x for x in s.dropna().astype(str)]))
    )
    for model, parts in tmp.items():
        if len(parts) >= 1:
            out[model] = parts[0]
        else:
            out[model] = model_part_ref.get(model)
    return out


def get_day_model_part_qty_map(demand: pd.DataFrame, current_date: pd.Timestamp) -> Dict[str, Dict[str, float]]:
    day = demand[demand["date"] == current_date].copy()
    out: Dict[str, Dict[str, float]] = {}
    if day.empty:
        return out
    grp = day.groupby(["MODEL_NO", "PART_NO"], as_index=False)["qty"].sum()
    for _, r in grp.iterrows():
        model = str(r["MODEL_NO"]).strip()
        part = str(r["PART_NO"]).strip()
        qty = float(r["qty"])
        out.setdefault(model, {})
        out[model][part] = out[model].get(part, 0.0) + qty
    return out


def assign_aoi_parts_to_lines(
    model: str,
    chosen_lines: List[Tuple],
    model_part_qty: Dict[str, float],
    line_state: Dict[str, Dict[str, Dict]],
) -> Dict[str, str]:
    if not chosen_lines:
        return {}

    norm_lines = []
    for item in chosen_lines:
        line = item[0]
        cap_day = float(item[1]) if len(item) > 1 and pd.notna(item[1]) else 0.0
        eff_cap = float(item[2]) if len(item) > 2 and pd.notna(item[2]) else cap_day
        norm_lines.append({
            "line": line,
            "cap_day": cap_day,
            "eff_cap": eff_cap,
        })

    if not model_part_qty:
        return {x["line"]: None for x in norm_lines}

    remaining = {
        str(p).strip(): float(q)
        for p, q in model_part_qty.items()
        if pd.notna(p) and float(q) > 1e-9
    }
    if not remaining:
        return {x["line"]: None for x in norm_lines}

    active_parts = sorted(remaining.items(), key=lambda kv: (-kv[1], kv[0]))
    assign: Dict[str, str] = {}
    unassigned = {x["line"]: x for x in norm_lines}

    def _line_rank_for_part(part: str, line_info: Dict) -> Tuple:
        prev = line_state["AOI"].get(line_info["line"], {})
        prev_model = prev.get("model")
        prev_part = prev.get("part")
        same_part_hist = int(prev_model == model and prev_part == part)
        same_model_hist = int(prev_model == model)
        return (-same_part_hist, -same_model_hist, -line_info["eff_cap"], -line_info["cap_day"], line_info["line"])

    def _best_line_for_part(part: str, lines_dict: Dict[str, Dict]) -> Optional[Dict]:
        if not lines_dict:
            return None
        ranked = sorted(lines_dict.values(), key=lambda li: _line_rank_for_part(part, li))
        return ranked[0] if ranked else None

    # 硬保留：同 model 多 part 时，历史上已跑该 model+part 的线，
    # 只要该 part 当天仍有需求/结转，就先锁给该 part，不参与后续平衡分配。
    hard_reserved_parts = set()
    for line_info in sorted(unassigned.values(), key=lambda li: li["line"]):
        line = line_info["line"]
        prev = line_state["AOI"].get(line, {})
        prev_model = prev.get("model")
        prev_part = prev.get("part")
        if prev_model == model and prev_part in remaining and remaining.get(prev_part, 0.0) > 1e-9:
            assign[line] = prev_part
            remaining[prev_part] = max(0.0, remaining[prev_part] - line_info["eff_cap"])
            hard_reserved_parts.add(prev_part)
    for line in list(assign.keys()):
        unassigned.pop(line, None)

    # 若剩余 line 数量足够，则未被硬保留覆盖到的 active part 至少先拿到 1 条可行线
    uncovered_parts = [(p, q) for p, q in active_parts if p not in hard_reserved_parts]
    if len(unassigned) >= len(uncovered_parts):
        for part, qty in uncovered_parts:
            if remaining.get(part, 0.0) <= 1e-9:
                continue
            best = _best_line_for_part(part, unassigned)
            if best is None:
                continue
            line = best["line"]
            assign[line] = part
            remaining[part] = max(0.0, remaining[part] - best["eff_cap"])
            unassigned.pop(line, None)

    # 剩余线继续按“剩余需求最大”分配；
    # 若某条线历史上就在跑某个仍有需求的 part，则优先延续该 part
    while unassigned:
        prev_pref = None
        for li in unassigned.values():
            prev = line_state["AOI"].get(li["line"], {})
            prev_part = prev.get("part")
            prev_model = prev.get("model")
            if prev_model == model and prev_part in remaining and remaining.get(prev_part, 0.0) > 1e-9:
                prev_pref = prev_part
                break

        available_parts = [(p, q) for p, q in remaining.items() if q > 1e-9]
        if prev_pref is not None:
            best_part = prev_pref
        elif available_parts:
            available_parts = sorted(available_parts, key=lambda kv: (-kv[1], kv[0]))
            best_part = available_parts[0][0]
        else:
            best_part = active_parts[0][0]

        best_line = _best_line_for_part(best_part, unassigned)
        if best_line is None:
            break
        line = best_line["line"]
        assign[line] = best_part
        remaining[best_part] = max(0.0, remaining.get(best_part, 0.0) - best_line["eff_cap"])
        unassigned.pop(line, None)

    return assign


# =========================================================
# 硬保留：只保留最少必要线数
# =========================================================

def build_cut_hard_reservations_for_day(
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    line_state: Dict[str, Dict[str, Dict]],
    cut_cap: pd.DataFrame,
    changeover: pd.DataFrame,
    hist_recipe: Optional[pd.DataFrame] = None,
    irregular_models: Optional[set] = None,
) -> Dict[str, set]:
    reservations = defaultdict(set)
    current_lines_by_key = defaultdict(list)

    for line, st in line_state["CUT"].items():
        key = st.get("key")
        model = st.get("model")
        part = st.get("part")
        if not key or not model:
            continue
        seg_len = consecutive_segment_len_product(demand, current_date, key, max_days=3)
        if seg_len <= 0:
            continue

        if not line_supports_full_cut_ratio_structure(cut_cap, key, line):
            continue
        cap_info = cut_cap[(cut_cap["key"] == key) & (cut_cap["line"] == line)]
        if cap_info.empty:
            continue
        cap_shift = float(cap_info["cap_shift"].max())
        cap_day = float(cap_info["cap_day"].max())
        first_chg_h = get_first_changeover_h_for_line(
            process="CUT",
            line=line,
            target_model=model,
            target_part=part,
            line_state=line_state,
            changeover_df=changeover,
            hist_recipe_df=hist_recipe,
            irregular_models=irregular_models,
        )
        eff_cap_day = get_effective_cap_day_after_first_changeover(cap_shift, first_chg_h)
        seg_cap = get_segment_capacity_contribution(eff_cap_day, cap_day, seg_len)
        current_lines_by_key[key].append((line, eff_cap_day, cap_day, seg_cap, first_chg_h))

    for key, lines in current_lines_by_key.items():
        seg_total = segment_total_product(demand, current_date, key, max_days=3)

        lines = sorted(lines, key=lambda x: (-x[3], x[4], -x[1], -x[2], x[0]))
        acc = 0.0
        for line, eff_cap_day, cap_day, seg_cap, first_chg_h in lines:
            reservations[key].add(line)
            acc += seg_cap
            if acc >= seg_total:
                break

    return reservations


def build_aoi_hard_reservations_for_day(
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    line_state: Dict[str, Dict[str, Dict]],
    aoi_cap: pd.DataFrame,
    changeover: pd.DataFrame,
    hist_recipe: Optional[pd.DataFrame] = None,
) -> Dict[str, set]:
    reservations = defaultdict(set)
    current_lines_by_model = defaultdict(list)

    for line, st in line_state["AOI"].items():
        model = st.get("model")
        part = st.get("part")
        if not model:
            continue
        seg_len = consecutive_segment_len_model(demand, current_date, model, max_days=3)
        if seg_len <= 0:
            continue

        cap_info = aoi_cap[(aoi_cap["MODEL_NO"] == model) & (aoi_cap["line"] == line)]
        if cap_info.empty:
            continue
        cap_shift = float(cap_info["cap_shift"].max())
        cap_day = float(cap_info["cap_day"].max())
        max_line = int(cap_info["MAX_LINE"].max())
        first_chg_h = get_first_changeover_h_for_line(
            process="AOI",
            line=line,
            target_model=model,
            target_part=part,
            line_state=line_state,
            changeover_df=changeover,
            hist_recipe_df=hist_recipe,
        )
        eff_cap_day = get_effective_cap_day_after_first_changeover(cap_shift, first_chg_h)
        seg_cap = get_segment_capacity_contribution(eff_cap_day, cap_day, seg_len)
        current_lines_by_model[model].append((line, eff_cap_day, cap_day, max_line, seg_cap, first_chg_h))

    for model, lines in current_lines_by_model.items():
        seg_total = segment_total_model(demand, current_date, model, max_days=3)

        lines = sorted(lines, key=lambda x: (-x[4], x[5], -x[1], -x[2], x[0]))
        max_line = max(x[3] for x in lines)

        acc = 0.0
        cnt = 0
        for line, eff_cap_day, cap_day, _, seg_cap, first_chg_h in lines:
            if cnt >= max_line:
                break
            reservations[model].add(line)
            acc += seg_cap
            cnt += 1
            if acc >= seg_total:
                break

    return reservations


# =========================================================
# 连续需求锁线：禁止预换出去
# =========================================================

def remaining_segment_days_for_product(demand: pd.DataFrame, current_date: pd.Timestamp, key: str, max_days: int = 3) -> int:
    return len(get_sticky_segment_dates_product(demand, current_date, key, max_days=max_days, gap_tolerance=1))


def is_line_locked_for_future_segment(
    process: str,
    line: str,
    line_state: Dict[str, Dict[str, Dict]],
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    cut_cap: pd.DataFrame,
    aoi_cap: pd.DataFrame,
) -> bool:
    st = line_state[process].get(line, {})
    if process == "CUT":
        key = st.get("key")
        if not key:
            return False
        future_days = remaining_segment_days_for_product(demand, current_date, key, max_days=3)
        if future_days <= 1:
            return False
        ok = ((cut_cap["line"] == line) & (cut_cap["key"] == key)).any()
        return bool(ok)

    if process == "AOI":
        model = st.get("model")
        if not model:
            return False
        future_days = consecutive_segment_len_model(demand, current_date, model, max_days=3)
        if future_days <= 1:
            return False
        ok = ((aoi_cap["line"] == line) & (aoi_cap["MODEL_NO"] == model)).any()
        return bool(ok)

    return False


# =========================================================
# Step1 分线
# =========================================================

def allocate_cut_lines_with_hard_reservation(
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    cut_cap: pd.DataFrame,
    line_state: Dict[str, Dict[str, Dict]],
    changeover: pd.DataFrame,
    hist_recipe: Optional[pd.DataFrame] = None,
    irregular_models: Optional[set] = None,
):
    day_demand = demand[demand["date"] == current_date].copy()
    if day_demand.empty:
        return pd.DataFrame(
            columns=["date", "process", "MODEL_NO", "PART_NO", "key", "line", "cap_day", "carryover", "planned_today_qty"]
        )

    products_today = day_demand[["MODEL_NO", "PART_NO", "key", "qty"]].copy()
    reservations = build_cut_hard_reservations_for_day(
        demand=demand,
        current_date=current_date,
        line_state=line_state,
        cut_cap=cut_cap,
        changeover=line_state["__meta__"]["changeover"],
        hist_recipe=line_state["__meta__"]["hist_recipe"],
        irregular_models=line_state["__meta__"]["irregular_models"],
    )

    reserved_lines_global = set()
    for _, lines in reservations.items():
        reserved_lines_global.update(lines)

    used_lines = set()
    cut_assign = []

    products_today = products_today.sort_values(["qty", "MODEL_NO", "PART_NO"], ascending=[False, True, True]).reset_index(drop=True)

    for _, row in products_today.iterrows():
        model = row["MODEL_NO"]
        part = row["PART_NO"]
        key = row["key"]
        today_qty = float(row["qty"])

        seg_len = consecutive_segment_len_product(demand, current_date, key, max_days=3)
        seg_total = segment_total_product(demand, current_date, key, max_days=3)

        cand = cut_cap[cut_cap["key"] == key].copy()
        if cand.empty:
            continue
        valid_lines_for_key = filter_cut_lines_supporting_full_ratio(cut_cap, key, sorted(cand["line"].dropna().astype(str).unique().tolist()))
        cand = cand[cand["line"].isin(valid_lines_for_key)].copy()
        if cand.empty:
            continue

        chosen = []
        reserved_for_key = reservations.get(key, set())

        # 关键修复：保留线一加入就立刻登记 used_lines，避免后面重复选中
        for line in reserved_for_key:
            if line in used_lines:
                continue
            cap_info = cand[cand["line"] == line]
            if cap_info.empty:
                continue
            cap_shift = float(cap_info["cap_shift"].max())
            cap_day = float(cap_info["cap_day"].max())
            first_chg_h = get_first_changeover_h_for_line(
                process="CUT",
                line=line,
                target_model=model,
                target_part=part,
                line_state=line_state,
                changeover_df=changeover,
                hist_recipe_df=hist_recipe,
                irregular_models=irregular_models,
            )
            eff_cap_day = get_effective_cap_day_after_first_changeover(cap_shift, first_chg_h)
            seg_cap = get_segment_capacity_contribution(eff_cap_day, cap_day, seg_len)
            chosen.append((line, cap_day, eff_cap_day, first_chg_h, seg_cap, 1))
            used_lines.add(line)

        current_cap_acc = sum(x[4] for x in chosen)

        model_lines = cand[["line", "cap_shift", "cap_day"]].drop_duplicates().copy()
        forbidden_lines = (reserved_lines_global - reserved_for_key) | used_lines
        model_lines = model_lines[~model_lines["line"].isin(forbidden_lines)].copy()
        if not model_lines.empty:
            model_lines["first_changeover_h"] = model_lines["line"].apply(
                lambda ln: get_first_changeover_h_for_line(
                    process="CUT",
                    line=ln,
                    target_model=model,
                    target_part=part,
                    line_state=line_state,
                    changeover_df=changeover,
                    hist_recipe_df=hist_recipe,
                    irregular_models=irregular_models,
                )
            )
            model_lines["effective_cap_day"] = model_lines.apply(
                lambda r: get_effective_cap_day_after_first_changeover(float(r["cap_shift"]), float(r["first_changeover_h"])), axis=1
            )
            model_lines["segment_cap"] = model_lines.apply(
                lambda r: get_segment_capacity_contribution(float(r["effective_cap_day"]), float(r["cap_day"]), seg_len), axis=1
            )
            model_lines = model_lines.sort_values(["segment_cap", "first_changeover_h", "effective_cap_day", "cap_day", "line"], ascending=[False, True, False, False, True])

        while current_cap_acc < seg_total and not model_lines.empty:
            r = model_lines.iloc[0]
            line = r["line"]
            cap_day = float(r["cap_day"])
            eff_cap_day = float(r["effective_cap_day"])
            first_chg_h = float(r["first_changeover_h"])
            seg_cap = float(r["segment_cap"])
            chosen.append((line, cap_day, eff_cap_day, first_chg_h, seg_cap, 0))
            used_lines.add(line)
            current_cap_acc += seg_cap
            model_lines = model_lines.iloc[1:]

        remain = today_qty
        for line, cap_day, eff_cap_day, first_chg_h, seg_cap, carry in chosen:
            alloc = min(eff_cap_day, remain)
            remain -= alloc
            if alloc <= 0:
                continue
            cut_assign.append({
                "date": current_date.date(),
                "process": "CUT",
                "MODEL_NO": model,
                "PART_NO": part,
                "key": key,
                "line": line,
                "cap_day": cap_day,
                "effective_cap_day": eff_cap_day,
                "first_changeover_h": first_chg_h,
                "carryover": carry,
                "planned_today_qty": alloc,
            })

    return pd.DataFrame(
        cut_assign,
        columns=["date", "process", "MODEL_NO", "PART_NO", "key", "line", "cap_day", "effective_cap_day", "first_changeover_h", "carryover", "planned_today_qty"]
    )


def allocate_aoi_lines_with_hard_reservation(
    demand: pd.DataFrame,
    current_date: pd.Timestamp,
    aoi_cap: pd.DataFrame,
    line_state: Dict[str, Dict[str, Dict]],
    changeover: pd.DataFrame,
    hist_recipe: Optional[pd.DataFrame] = None,
):
    day_demand = demand[demand["date"] == current_date].copy()
    prev_carry_pool = line_state.get("__meta__", {}).get("aoi_prev_carry_pool", {}) or {}

    # AOI 要同时考虑：
    # 1) 当天原始需求
    # 2) 前一生产日结转的WIP
    model_part_qty_map: Dict[str, Dict[str, float]] = defaultdict(dict)

    if not day_demand.empty:
        day_part_map = get_day_model_part_qty_map(demand, current_date)
        for model, part_map in day_part_map.items():
            for part, qty in part_map.items():
                model_part_qty_map[model][part] = model_part_qty_map[model].get(part, 0.0) + float(qty)

    for (model, part), qty in prev_carry_pool.items():
        if float(qty) > 1e-9:
            model_part_qty_map[model][part] = model_part_qty_map[model].get(part, 0.0) + float(qty)

    if not model_part_qty_map:
        return pd.DataFrame(
            columns=["date", "process", "MODEL_NO", "PART_NO", "line", "cap_day", "carryover", "planned_today_qty"]
        )

    models_today = pd.DataFrame([
        {"MODEL_NO": model, "today_qty": float(sum(part_map.values()))}
        for model, part_map in model_part_qty_map.items()
    ])

    reservations = build_aoi_hard_reservations_for_day(
        demand=demand,
        current_date=current_date,
        line_state=line_state,
        aoi_cap=aoi_cap,
        changeover=line_state["__meta__"]["changeover"],
        hist_recipe=line_state["__meta__"]["hist_recipe"],
    )

    reserved_lines_global = set()
    for _, lines in reservations.items():
        reserved_lines_global.update(lines)

    aoi_assign = []
    used_lines = set()

    models_today = models_today.sort_values(["today_qty", "MODEL_NO"], ascending=[False, True]).reset_index(drop=True)

    for _, row in models_today.iterrows():
        model = row["MODEL_NO"]
        today_qty = float(row["today_qty"])
        part_qty_map = dict(model_part_qty_map.get(model, {}))
        part = sorted(part_qty_map.items(), key=lambda kv: (-kv[1], kv[0]))[0][0] if part_qty_map else None

        seg_len = consecutive_segment_len_model(demand, current_date, model, max_days=3)
        if seg_len <= 0:
            # carry-only model 也至少按当天处理 1 天来选线
            seg_len = 1
        seg_total = max(today_qty, segment_total_model(demand, current_date, model, max_days=3))

        cand = aoi_cap[aoi_cap["MODEL_NO"] == model].copy()
        if cand.empty:
            continue

        chosen = []
        max_line = int(cand["MAX_LINE"].max())
        active_parts = [(p, q) for p, q in sorted(part_qty_map.items(), key=lambda kv: (-kv[1], kv[0])) if q > 1e-9]

        # 先锁定“历史同 model 同 part”的线：若该 part 今天有需求/结转，则尽量保住至少1条
        for apart, _qty in active_parts:
            if len(chosen) >= max_line:
                break
            same_part_cands = []
            for line, st in line_state["AOI"].items():
                if line in used_lines:
                    continue
                if st.get("model") != model or st.get("part") != apart:
                    continue
                cap_info = cand[cand["line"] == line]
                if cap_info.empty:
                    continue
                cap_shift = float(cap_info["cap_shift"].max())
                cap_day = float(cap_info["cap_day"].max())
                first_chg_h = get_first_changeover_h_for_line(
                    process="AOI",
                    line=line,
                    target_model=model,
                    target_part=apart,
                    line_state=line_state,
                    changeover_df=changeover,
                    hist_recipe_df=hist_recipe,
                )
                eff_cap_day = get_effective_cap_day_after_first_changeover(cap_shift, first_chg_h)
                seg_cap = get_segment_capacity_contribution(eff_cap_day, cap_day, seg_len)
                same_part_cands.append((line, cap_day, eff_cap_day, first_chg_h, seg_cap, 1))
            if same_part_cands:
                same_part_cands = sorted(same_part_cands, key=lambda x: (x[3], -x[2], -x[1], x[0]))
                best = same_part_cands[0]
                if best[0] not in used_lines:
                    chosen.append(best)
                    used_lines.add(best[0])

        reserved_for_model = set(reservations.get(model, set()))
        reserved_for_model.update([x[0] for x in chosen])

        # 其余 reservation 继续补入
        for line in sorted(reserved_for_model):
            if len(chosen) >= max_line:
                break
            if line in used_lines:
                continue
            cap_info = cand[cand["line"] == line]
            if cap_info.empty:
                continue
            cap_shift = float(cap_info["cap_shift"].max())
            cap_day = float(cap_info["cap_day"].max())
            first_chg_h = get_first_changeover_h_for_line(
                process="AOI",
                line=line,
                target_model=model,
                target_part=part,
                line_state=line_state,
                changeover_df=changeover,
                hist_recipe_df=hist_recipe,
            )
            eff_cap_day = get_effective_cap_day_after_first_changeover(cap_shift, first_chg_h)
            seg_cap = get_segment_capacity_contribution(eff_cap_day, cap_day, seg_len)
            chosen.append((line, cap_day, eff_cap_day, first_chg_h, seg_cap, 1))
            used_lines.add(line)

        current_cap_acc = sum(x[4] for x in chosen)

        model_lines = cand[["line", "cap_shift", "cap_day"]].drop_duplicates().copy()
        forbidden_lines = (reserved_lines_global - reservations.get(model, set())) | used_lines
        model_lines = model_lines[~model_lines["line"].isin(forbidden_lines)].copy()
        if not model_lines.empty:
            model_lines["first_changeover_h"] = model_lines["line"].apply(
                lambda ln: get_first_changeover_h_for_line(
                    process="AOI",
                    line=ln,
                    target_model=model,
                    target_part=part,
                    line_state=line_state,
                    changeover_df=changeover,
                    hist_recipe_df=hist_recipe,
                )
            )
            model_lines["effective_cap_day"] = model_lines.apply(
                lambda r: get_effective_cap_day_after_first_changeover(float(r["cap_shift"]), float(r["first_changeover_h"])), axis=1
            )
            model_lines["segment_cap"] = model_lines.apply(
                lambda r: get_segment_capacity_contribution(float(r["effective_cap_day"]), float(r["cap_day"]), seg_len), axis=1
            )
            model_lines = model_lines.sort_values(
                ["segment_cap", "first_changeover_h", "effective_cap_day", "cap_day", "line"],
                ascending=[False, True, False, False, True]
            )

        min_lines_for_parts = min(max_line, max(1, len(active_parts))) if active_parts else 1
        while len(chosen) < max_line and not model_lines.empty:
            # 先保证 active parts 至少各有一条线（如果 max_line 允许）
            if len(chosen) >= min_lines_for_parts and current_cap_acc >= seg_total:
                break
            r = model_lines.iloc[0]
            line = r["line"]
            cap_day = float(r["cap_day"])
            eff_cap_day = float(r["effective_cap_day"])
            first_chg_h = float(r["first_changeover_h"])
            seg_cap = float(r["segment_cap"])
            chosen.append((line, cap_day, eff_cap_day, first_chg_h, seg_cap, 0))
            used_lines.add(line)
            current_cap_acc += seg_cap
            model_lines = model_lines.iloc[1:]

        line_part_map = assign_aoi_parts_to_lines(model, chosen, part_qty_map, line_state)

        remain = today_qty
        for line, cap_day, eff_cap_day, first_chg_h, seg_cap, carry in chosen:
            alloc = min(eff_cap_day, remain)
            remain -= alloc
            if alloc <= 0 and eff_cap_day <= 0:
                continue
            aoi_assign.append({
                "date": current_date.date(),
                "process": "AOI",
                "MODEL_NO": model,
                "PART_NO": line_part_map.get(line, part),
                "line": line,
                "cap_day": cap_day,
                "effective_cap_day": eff_cap_day,
                "first_changeover_h": first_chg_h,
                "carryover": carry,
                "planned_today_qty": max(alloc, 0.0),
            })

    return pd.DataFrame(
        aoi_assign,
        columns=["date", "process", "MODEL_NO", "PART_NO", "line", "cap_day", "effective_cap_day", "first_changeover_h", "carryover", "planned_today_qty"]
    )


def allocate_lines_step1_for_day(
    data: Dict,
    current_date: pd.Timestamp,
    line_state: Dict[str, Dict[str, Dict]],
):
    demand = data["demand"]
    day_demand = demand[demand["date"] == current_date].copy()

    step1_product = []

    cut_assign_df = allocate_cut_lines_with_hard_reservation(
        demand=demand,
        current_date=current_date,
        cut_cap=data["cut_cap"],
        line_state=line_state,
        changeover=data["changeover"],
        hist_recipe=data["hist_recipe"],
        irregular_models=data["irregular_models"],
    )

    if not cut_assign_df.empty:
        cut_summary = (
            cut_assign_df.groupby(["MODEL_NO", "PART_NO", "key"], as_index=False)
            .agg(
                assigned_lines=("line", "nunique"),
                carryover_kept=("carryover", "sum"),
                planned_today_qty=("planned_today_qty", "sum"),
            )
        )
        cut_summary = cut_summary.merge(
            day_demand[["MODEL_NO", "PART_NO", "key", "qty"]].rename(columns={"qty": "today_demand"}),
            on=["MODEL_NO", "PART_NO", "key"],
            how="left",
        )

        for _, rr in cut_summary.iterrows():
            model = rr["MODEL_NO"]
            part = rr["PART_NO"]
            key = rr["key"]
            seg_len = consecutive_segment_len_product(demand, current_date, key, max_days=3)
            seg_total = segment_total_product(demand, current_date, key, max_days=3)

            step1_product.append({
                "date": current_date.date(),
                "process": "CUT",
                "MODEL_NO": model,
                "PART_NO": part,
                "today_demand": float(rr["today_demand"]),
                "segment_len": seg_len,
                "deadline": get_segment_deadline_product(demand, current_date, key, max_days=3).date(),
                "segment_total": seg_total,
                "assigned_lines": int(rr["assigned_lines"]),
                "carryover_kept": int(rr["carryover_kept"]),
                "planned_today_qty": float(rr["planned_today_qty"]),
                "today_gap": max(0.0, float(rr["today_demand"]) - float(rr["planned_today_qty"])),
            })

    aoi_assign_df = allocate_aoi_lines_with_hard_reservation(
        demand=demand,
        current_date=current_date,
        aoi_cap=data["aoi_cap"],
        line_state=line_state,
        changeover=data["changeover"],
        hist_recipe=data["hist_recipe"],
    )

    if not aoi_assign_df.empty:
        aoi_summary = (
            aoi_assign_df.groupby(["MODEL_NO", "PART_NO"], as_index=False)
            .agg(
                assigned_lines=("line", "nunique"),
                carryover_kept=("carryover", "sum"),
                planned_today_qty=("planned_today_qty", "sum"),
            )
        )

        model_day_demand = (
            day_demand.groupby(["MODEL_NO", "PART_NO"], as_index=False)["qty"]
            .sum()
            .rename(columns={"qty": "today_demand"})
        )
        aoi_summary = aoi_summary.merge(model_day_demand, on=["MODEL_NO", "PART_NO"], how="left")

        for _, rr in aoi_summary.iterrows():
            model = rr["MODEL_NO"]
            part = rr["PART_NO"]
            seg_len = consecutive_segment_len_model(demand, current_date, model, max_days=3)
            seg_total = segment_total_model(demand, current_date, model, max_days=3)

            step1_product.append({
                "date": current_date.date(),
                "process": "AOI",
                "MODEL_NO": model,
                "PART_NO": part,
                "today_demand": float(rr["today_demand"]),
                "segment_len": seg_len,
                "deadline": get_segment_deadline_model(demand, current_date, model, max_days=3).date(),
                "segment_total": seg_total,
                "assigned_lines": int(rr["assigned_lines"]),
                "carryover_kept": int(rr["carryover_kept"]),
                "planned_today_qty": float(rr["planned_today_qty"]),
                "today_gap": max(0.0, float(rr["today_demand"]) - float(rr["planned_today_qty"])),
            })

    step1_product_df = pd.DataFrame(
        step1_product,
        columns=[
            "date", "process", "MODEL_NO", "PART_NO", "today_demand",
            "segment_len", "deadline", "segment_total",
            "assigned_lines", "carryover_kept",
            "planned_today_qty", "today_gap"
        ],
    )

    return step1_product_df, cut_assign_df, aoi_assign_df


# =========================================================
# 汇总表
# =========================================================

def build_master_schedule_table(
    ops_df: pd.DataFrame,
    backlog_df: pd.DataFrame,
    warnings_df: pd.DataFrame,
) -> pd.DataFrame:
    if ops_df is None or ops_df.empty:
        return pd.DataFrame(columns=[
            "date", "process", "MODEL_NO", "PART_NO", "key", "line",
            "day_qty", "night_qty", "total_qty", "cap_shift", "changeover_h",
            "orig_demand", "has_carry_in", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet", "warning_info"
        ])

    ops = ops_df.copy()
    backlog = backlog_df.copy() if backlog_df is not None else pd.DataFrame()
    warns = warnings_df.copy() if warnings_df is not None else pd.DataFrame()

    for col in ["PART_NO", "key"]:
        if col not in ops.columns:
            ops[col] = None

    if backlog.empty:
        backlog = pd.DataFrame(columns=[
            "date", "process", "MODEL_NO", "PART_NO",
            "orig_demand", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet"
        ])
    else:
        if "PART_NO" not in backlog.columns:
            backlog["PART_NO"] = None
        for c in ["same_day_cut_in", "available_wip"]:
            if c not in backlog.columns:
                backlog[c] = 0.0
        backlog = backlog[
            ["date", "process", "MODEL_NO", "PART_NO", "orig_demand", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet"]
        ].copy()

    if warns.empty:
        warn_model = pd.DataFrame(columns=["date", "process", "MODEL_NO", "PART_NO", "warning_info"])
        warn_line = pd.DataFrame(columns=["date", "process", "MODEL_NO", "PART_NO", "line", "warning_info_line"])
    else:
        warns = warns.rename(columns={"model": "MODEL_NO", "part": "PART_NO"}).copy()
        if "PART_NO" not in warns.columns:
            warns["PART_NO"] = None

        warn_model = (
            warns.groupby(["date", "process", "MODEL_NO", "PART_NO"], dropna=False)["msg"]
            .apply(lambda s: " | ".join(pd.unique(s.astype(str))))
            .reset_index(name="warning_info")
        )

        warn_line = (
            warns[warns["line"].notna()]
            .groupby(["date", "process", "MODEL_NO", "PART_NO", "line"], dropna=False)["msg"]
            .apply(lambda s: " | ".join(pd.unique(s.astype(str))))
            .reset_index(name="warning_info_line")
        )

    master = ops.merge(
        backlog,
        on=["date", "process", "MODEL_NO", "PART_NO"],
        how="left"
    )

    master = master.merge(
        warn_model,
        on=["date", "process", "MODEL_NO", "PART_NO"],
        how="left"
    )

    master = master.merge(
        warn_line,
        on=["date", "process", "MODEL_NO", "PART_NO", "line"],
        how="left"
    )

    master["orig_demand"] = master["orig_demand"].fillna(0.0)
    master["carry_in"] = master["carry_in"].fillna(0.0)
    if "same_day_cut_in" not in master.columns:
        master["same_day_cut_in"] = 0.0
    if "available_wip" not in master.columns:
        master["available_wip"] = 0.0
    master["same_day_cut_in"] = master["same_day_cut_in"].fillna(0.0)
    master["available_wip"] = master["available_wip"].fillna(0.0)
    master["carry_out"] = master["carry_out"].fillna(0.0)
    master["unmet"] = master["unmet"].fillna(0.0)
    master["has_carry_in"] = master["carry_in"] > 0

    def combine_warning(r):
        a = r.get("warning_info")
        b = r.get("warning_info_line")
        if pd.notna(a) and pd.notna(b):
            if str(b) in str(a):
                return a
            return str(a) + " | " + str(b)
        if pd.notna(a):
            return a
        if pd.notna(b):
            return b
        return None

    master["warning_info"] = master.apply(combine_warning, axis=1)

    keep_cols = [
        "date", "process", "MODEL_NO", "PART_NO", "key", "line",
        "day_qty", "night_qty", "total_qty", "cap_shift", "changeover_h",
        "orig_demand", "has_carry_in", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet", "warning_info"
    ]
    for c in keep_cols:
        if c not in master.columns:
            master[c] = None

    return master[keep_cols].sort_values(
        ["date", "process", "MODEL_NO", "PART_NO", "line"],
        na_position="last"
    )


def build_all_changeovers_table(changes_df: pd.DataFrame, prechg_df: pd.DataFrame) -> pd.DataFrame:
    chg = changes_df.copy() if changes_df is not None else pd.DataFrame()
    pre = prechg_df.copy() if prechg_df is not None else pd.DataFrame()

    if chg.empty:
        chg = pd.DataFrame(columns=[
            "date", "process", "line", "from_model", "from_part",
            "to_model", "to_part", "changeover_h", "shift"
        ])
    else:
        chg = chg.copy()
        chg["change_type"] = "production_changeover"

    if pre.empty:
        pre = pd.DataFrame(columns=[
            "date", "process", "line", "from_model", "from_part",
            "to_model", "to_part", "changeover_h", "for_next_day", "planned"
        ])
    else:
        pre = pre.rename(columns={"free_time_today": "free_time_today_h"}).copy()
        pre["shift"] = "prechangeover_after_schedule"
        pre["change_type"] = "pre_changeover"

    keep_chg = [
        "date", "process", "line", "from_model", "from_part",
        "to_model", "to_part", "changeover_h", "shift", "change_type"
    ]
    keep_pre = [
        "date", "process", "line", "from_model", "from_part",
        "to_model", "to_part", "changeover_h", "shift", "change_type"
    ]

    for c in keep_chg:
        if c not in chg.columns:
            chg[c] = None
    for c in keep_pre:
        if c not in pre.columns:
            pre[c] = None

    out = pd.concat([chg[keep_chg], pre[keep_pre]], ignore_index=True)
    return out.sort_values(["date", "process", "line", "change_type"], na_position="last")




def build_master_schedule_detail_table(detail_df: pd.DataFrame, backlog_df: pd.DataFrame, warnings_df: pd.DataFrame) -> pd.DataFrame:
    if detail_df is None or detail_df.empty:
        return pd.DataFrame(columns=[
            "date", "process", "detail_type", "segment_label", "MODEL_NO", "PART_NO", "key", "line",
            "ratio_pct", "changeover_h", "day_qty", "night_qty", "total_qty",
            "orig_demand", "has_carry_in", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet", "warning_info"
        ])

    detail = detail_df.copy()
    backlog = backlog_df.copy() if backlog_df is not None else pd.DataFrame()
    warns = warnings_df.copy() if warnings_df is not None else pd.DataFrame()

    if backlog.empty:
        backlog = pd.DataFrame(columns=["date", "process", "MODEL_NO", "PART_NO", "orig_demand", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet"])
    else:
        if "PART_NO" not in backlog.columns:
            backlog["PART_NO"] = None
        for c in ["same_day_cut_in", "available_wip"]:
            if c not in backlog.columns:
                backlog[c] = 0.0
        backlog = backlog[["date", "process", "MODEL_NO", "PART_NO", "orig_demand", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet"]].copy()

    if warns.empty:
        warn_model = pd.DataFrame(columns=["date", "process", "MODEL_NO", "PART_NO", "warning_info"])
        warn_line = pd.DataFrame(columns=["date", "process", "MODEL_NO", "PART_NO", "line", "warning_info_line"])
    else:
        warns = warns.rename(columns={"model": "MODEL_NO", "part": "PART_NO"}).copy()
        if "PART_NO" not in warns.columns:
            warns["PART_NO"] = None

        warn_model = (
            warns.groupby(["date", "process", "MODEL_NO", "PART_NO"], dropna=False)["msg"]
            .apply(lambda s: " | ".join(pd.unique(s.astype(str))))
            .reset_index(name="warning_info")
        )
        warn_line = (
            warns[warns["line"].notna()]
            .groupby(["date", "process", "MODEL_NO", "PART_NO", "line"], dropna=False)["msg"]
            .apply(lambda s: " | ".join(pd.unique(s.astype(str))))
            .reset_index(name="warning_info_line")
        )

    out = detail.merge(backlog, on=["date", "process", "MODEL_NO", "PART_NO"], how="left", suffixes=("_detail", ""))
    out = out.merge(warn_model, on=["date", "process", "MODEL_NO", "PART_NO"], how="left")
    out = out.merge(warn_line, on=["date", "process", "MODEL_NO", "PART_NO", "line"], how="left")

    for c in ["orig_demand", "carry_in"]:
        detail_c = f"{c}_detail"
        if detail_c in out.columns:
            if c not in out.columns:
                out[c] = out[detail_c]
            else:
                out[c] = out[detail_c].where(out[detail_c].notna(), out[c])
            out = out.drop(columns=[detail_c])

    for c in ["orig_demand", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet"]:
        if c not in out.columns:
            out[c] = 0.0
        out[c] = out[c].fillna(0.0)
    out["has_carry_in"] = out["carry_in"] > 0

    def combine_warning(r):
        a = r.get("warning_info")
        b = r.get("warning_info_line")
        if pd.notna(a) and pd.notna(b):
            if str(b) in str(a):
                return a
            return str(a) + " | " + str(b)
        if pd.notna(a):
            return a
        if pd.notna(b):
            return b
        return None

    out["warning_info"] = out.apply(combine_warning, axis=1)

    keep_cols = [
        "date", "process", "detail_type", "event_seq", "segment_no", "segment_label",
        "MODEL_NO", "PART_NO", "key", "line", "ratio_pct", "ratio_frac",
        "from_model", "from_part", "to_model", "to_part", "shift",
        "changeover_h", "day_qty", "night_qty", "total_qty", "cap_shift", "target_qty",
        "orig_demand", "has_carry_in", "carry_in", "same_day_cut_in", "available_wip", "carry_out", "unmet", "warning_info"
    ]
    for c in keep_cols:
        if c not in out.columns:
            out[c] = None
    return out[keep_cols].sort_values(["date", "process", "MODEL_NO", "PART_NO", "line", "event_seq"], na_position="last").reset_index(drop=True)


def export_results_to_excel(output_path: str, sheets: Dict[str, pd.DataFrame]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import math
    import numpy as np

    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin)
    alt_fill = PatternFill("solid", fgColor="F7FBFF")
    note_fill = PatternFill("solid", fgColor="FFF2CC")

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if df is None:
            df = pd.DataFrame()
        if df.empty:
            ws["A1"] = "No data"
            continue

        out = df.copy()
        for col in out.columns:
            if pd.api.types.is_datetime64_any_dtype(out[col]):
                out[col] = out[col].dt.strftime("%Y-%m-%d %H:%M:%S")

        for j, col in enumerate(out.columns, 1):
            cell = ws.cell(row=1, column=j, value=str(col))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for i, row in enumerate(out.itertuples(index=False), 2):
            for j, val in enumerate(row, 1):
                if isinstance(val, (np.floating, float)) and (math.isnan(val) if isinstance(val, float) else pd.isna(val)):
                    val = None
                elif pd.isna(val):
                    val = None
                ws.cell(row=i, column=j, value=val)
            if i % 2 == 0 and sheet_name not in ("summary", "logic_check"):
                for j in range(1, len(out.columns) + 1):
                    ws.cell(row=i, column=j).fill = alt_fill

        ws.freeze_panes = "A2"
        #ws.auto_filter.ref = ws.dimensions

        try:
            ref = ws.dimensions
            table_name = ("T_" + "".join(ch for ch in sheet_name if ch.isalnum()))[:25]
            tab = Table(displayName=table_name, ref=ref)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)
        except Exception:
            pass

        for idx, col in enumerate(out.columns, 1):
            sample = [str(col)] + ["" if x is None else str(x) for x in out[col].head(200).tolist()]
            width = min(max(len(s) for s in sample) + 2, 45)
            if sheet_name in ("warnings", "logic_check"):
                width = min(max(width, 18), 60)
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.row_dimensions[1].height = 22

    if "summary" in wb.sheetnames:
        ws = wb["summary"]
        ws["A4"] = "Note"
        ws["B4"] = "master_schedule_detail / cut_detail 已拆分 CUT 第1段、WITH_RATIO、第2段（如有）"
        ws["A4"].fill = note_fill
        ws["B4"].fill = note_fill

    wb.save(output_path)


# =========================================================
# Step2 主体
# =========================================================

def init_line_state_from_hist(hist: pd.DataFrame) -> Dict[str, Dict[str, Dict]]:
    line_state = {"CUT": {}, "AOI": {}, "__meta__": {}}
    for _, r in hist.iterrows():
        process = r["PROCESS"]
        line = r["LINE"]
        model = r["MODEL_NO"]
        part = r["PART_NO"]
        key = r["key"]
        is_idle = key is None or model is None
        line_state[process][line] = {
            "model": model,
            "part": part,
            "key": key,
            "status": "IDLE" if is_idle else "RUNNING",
        }
    return line_state


def daily_changeover_limits(process: str) -> Dict[str, int]:
    if process == "CUT":
        return {"day": 4, "night": 2}
    if process == "AOI":
        return {"day": 2, "night": 0}
    raise ValueError(process)


def plan_shift_qty_with_changeover_shift(cap_shift: float, changeover_h: float, changeover_shift: str) -> Tuple[float, float]:
    if changeover_shift == "day":
        day_prod_h = max(0.0, 12.0 - changeover_h)
        night_prod_h = max(0.0, 24.0 - max(12.0, changeover_h)) if changeover_h > 12.0 else 12.0
        day_qty = cap_shift * day_prod_h / 12.0
        night_qty = cap_shift * night_prod_h / 12.0
    elif changeover_shift == "night":
        # 目标机种从夜班开始切换，因此白班不生产目标机种；若换线超过12小时，则夜班也只生产剩余可用时长
        day_qty = 0.0
        night_prod_h = max(0.0, 12.0 - changeover_h)
        night_qty = cap_shift * night_prod_h / 12.0
    else:
        day_qty = cap_shift
        night_qty = cap_shift
    return day_qty, night_qty


def try_prechangeover(
    current_date: pd.Timestamp,
    next_date: pd.Timestamp,
    process: str,
    line_state: Dict[str, Dict[str, Dict]],
    free_time_today_by_line: Dict[str, float],
    remaining_changeovers_today: int,
    demand: pd.DataFrame,
    cut_cap: pd.DataFrame,
    aoi_cap: pd.DataFrame,
    changeover: pd.DataFrame,
    hist_recipe: Optional[pd.DataFrame] = None,
    irregular_models: Optional[set] = None,
):
    suggestions = []
    if remaining_changeovers_today <= 0:
        return suggestions

    if process == "CUT":
        next_day_demand = demand[demand["date"] == next_date].copy()
        next_targets = next_day_demand[["MODEL_NO", "PART_NO", "key"]].drop_duplicates().to_dict("records")

        for line, st in line_state["CUT"].items():
            if remaining_changeovers_today <= 0:
                break
            if free_time_today_by_line.get(line, 0.0) <= 0:
                continue

            current_key = st.get("key")
            current_model = st.get("model")
            current_part = st.get("part")

            if is_line_locked_for_future_segment(
                process="CUT",
                line=line,
                line_state=line_state,
                demand=demand,
                current_date=current_date,
                cut_cap=cut_cap,
                aoi_cap=aoi_cap,
            ):
                continue

            if current_key is not None and next_day_has_same_product(demand, current_date, current_key):
                continue

            best = None
            for tg in next_targets:
                cap_match = cut_cap[(cut_cap["line"] == line) & (cut_cap["key"] == tg["key"])]
                if cap_match.empty:
                    continue
                ratio_is_100 = bool(cap_match["ratio_is_100"].max()) if "ratio_is_100" in cap_match.columns else None
                chg_h = get_changeover_h(
                    "CUT",
                    current_model,
                    current_part,
                    tg["MODEL_NO"],
                    tg["PART_NO"],
                    changeover,
                    line=line,
                    hist_recipe_df=hist_recipe,
                    irregular_models=irregular_models,
                    ratio_is_100=ratio_is_100,
                )
                free_h = free_time_today_by_line.get(line, 0.0)
                if chg_h <= free_h or chg_h > 12.0:
                    best = (tg, chg_h)
                    break

            if best is not None:
                tg, chg_h = best
                suggestions.append({
                    "date": current_date.date(),
                    "for_next_day": next_date.date(),
                    "process": "CUT",
                    "line": line,
                    "from_model": current_model,
                    "from_part": current_part,
                    "to_model": tg["MODEL_NO"],
                    "to_part": tg["PART_NO"],
                    "changeover_h": chg_h,
                    "free_time_today": free_time_today_by_line.get(line, 0.0),
                    "planned": True,
                })
                line_state["CUT"][line] = {
                    "model": tg["MODEL_NO"],
                    "part": tg["PART_NO"],
                    "key": tg["key"],
                }
                remaining_changeovers_today -= 1

    elif process == "AOI":
        next_day_demand = demand[demand["date"] == next_date].copy()
        next_targets = next_day_demand[["MODEL_NO", "PART_NO"]].drop_duplicates().to_dict("records")

        for line, st in line_state["AOI"].items():
            if remaining_changeovers_today <= 0:
                break
            if free_time_today_by_line.get(line, 0.0) <= 0:
                continue

            current_model = st.get("model")
            current_part = st.get("part")

            if is_line_locked_for_future_segment(
                process="AOI",
                line=line,
                line_state=line_state,
                demand=demand,
                current_date=current_date,
                cut_cap=cut_cap,
                aoi_cap=aoi_cap,
            ):
                continue

            if current_model is not None and next_day_has_same_model(demand, current_date, current_model):
                continue

            best = None
            for tg in next_targets:
                ok = ((aoi_cap["line"] == line) & (aoi_cap["MODEL_NO"] == tg["MODEL_NO"])).any()
                if not ok:
                    continue
                chg_h = get_changeover_h(
                    "AOI",
                    current_model,
                    current_part,
                    tg["MODEL_NO"],
                    tg["PART_NO"],
                    changeover,
                    line=line,
                    hist_recipe_df=hist_recipe,
                    irregular_models=irregular_models,
                    ratio_is_100=None,
                )
                free_h = free_time_today_by_line.get(line, 0.0)
                if chg_h <= free_h or chg_h > 12.0:
                    best = (tg, chg_h)
                    break

            if best is not None:
                tg, chg_h = best
                suggestions.append({
                    "date": current_date.date(),
                    "for_next_day": next_date.date(),
                    "process": "AOI",
                    "line": line,
                    "from_model": current_model,
                    "from_part": current_part,
                    "to_model": tg["MODEL_NO"],
                    "to_part": tg["PART_NO"],
                    "changeover_h": chg_h,
                    "free_time_today": free_time_today_by_line.get(line, 0.0),
                    "planned": True,
                })
                line_state["AOI"][line] = {
                    "model": tg["MODEL_NO"],
                    "part": tg["PART_NO"],
                    "key": None,
                }
                remaining_changeovers_today -= 1

    return suggestions



def run_step2_with_backlog(
    data: Dict,
    start_date: str = "2026-02-04",
    end_date: str = "2026-02-19",
):
    demand = data["demand"]
    hist = data["hist"]
    cut_cap = data["cut_cap"]
    aoi_cap = data["aoi_cap"]
    changeover = data["changeover"]
    hist_recipe = data.get("hist_recipe")
    irregular_models = data.get("irregular_models", set())
    model_part_ref = build_model_part_reference(demand)

    line_state = init_line_state_from_hist(hist)
    line_state["__meta__"] = {
        "changeover": changeover,
        "hist_recipe": hist_recipe,
        "irregular_models": irregular_models,
    }
    demand.attrs["calendar_start"] = pd.Timestamp(start_date).normalize()
    demand.attrs["calendar_end"] = pd.Timestamp(end_date).normalize()
    dates = get_production_dates(demand)

    daily_rows = []
    op_rows = []
    detail_rows = []
    chg_rows = []
    warn_rows = []
    backlog_rows = []
    prechg_rows = []

    cut_backlog = defaultdict(float)   # key -> qty
    aoi_wip_pool = defaultdict(float)  # (model, part) -> qty 真实WIP池

    for d in dates:
        line_state["__meta__"]["aoi_prev_carry_pool"] = dict(aoi_wip_pool)
        day_model_part_map = get_day_model_part_map(demand, d, model_part_ref)
        step1_prod, cut_assign, aoi_assign = allocate_lines_step1_for_day(data, d, line_state)

        if cut_assign is None or cut_assign.empty:
            cut_assign = pd.DataFrame(
                columns=["date", "process", "MODEL_NO", "PART_NO", "key", "line", "cap_day", "carryover", "planned_today_qty"]
            )
        if aoi_assign is None or aoi_assign.empty:
            aoi_assign = pd.DataFrame(
                columns=["date", "process", "MODEL_NO", "PART_NO", "line", "cap_day", "carryover", "planned_today_qty"]
            )

        # =========================
        # CUT 班次排产
        # =========================
        cut_limits = daily_changeover_limits("CUT")
        cut_day_chg_used = 0
        cut_night_chg_used = 0
        free_time_cut = defaultdict(float)
        day_cut_actual_by_key = defaultdict(float)

        day_cut_demand_map = defaultdict(float)
        day_prod_map = {}

        raw_cut_today = demand[demand["date"] == d][["MODEL_NO", "PART_NO", "key", "qty"]].copy()
        for _, r in raw_cut_today.iterrows():
            day_cut_demand_map[r["key"]] += float(r["qty"])
            day_prod_map[r["key"]] = (r["MODEL_NO"], r["PART_NO"])

        for key, carry in list(cut_backlog.items()):
            day_cut_demand_map[key] += float(carry)
            if key not in day_prod_map:
                model, part = key.split("|", 1)
                day_prod_map[key] = (model, part)

        assigned_by_key = {}
        if not cut_assign.empty and "key" in cut_assign.columns:
            assigned_by_key = {k: v for k, v in cut_assign.groupby("key")}

        all_cut_keys_today = list(day_cut_demand_map.keys())

        for key in all_cut_keys_today:
            if key not in assigned_by_key:
                model, part = day_prod_map[key]
                cand = cut_cap[cut_cap["key"] == key].copy()
                if cand.empty:
                    warn_rows.append({
                        "date": d.date(),
                        "process": "CUT",
                        "model": model,
                        "part": part,
                        "line": None,
                        "warning_type": "cut_no_capability",
                        "msg": f"无CUT capability，需求/结转无法生产，缺口 {day_cut_demand_map[key]:.2f}",
                    })
                    continue

                valid_lines_for_key = filter_cut_lines_supporting_full_ratio(cut_cap, key, sorted(cand["line"].dropna().astype(str).unique().tolist()))
                cand = cand[cand["line"].isin(valid_lines_for_key)].copy()
                if cand.empty:
                    warn_rows.append({
                        "date": d.date(),
                        "process": "CUT",
                        "model": model,
                        "part": part,
                        "line": None,
                        "warning_type": "cut_no_full_ratio_line",
                        "msg": f"该CUT产品存在多ratio，但没有任何line支持完整ratio结构，需求/结转无法生产，缺口 {day_cut_demand_map[key]:.2f}",
                    })
                    continue

                cand_eval = []
                for line, g in cand.groupby("line"):
                    cap_shift = float(g["cap_shift"].max())
                    raw_cap_day = float(g["cap_day"].max())
                    first_chg_h = get_first_changeover_h_for_line(
                        process="CUT",
                        line=line,
                        target_model=model,
                        target_part=part,
                        line_state=line_state,
                        changeover_df=changeover,
                        hist_recipe_df=hist_recipe,
                        irregular_models=irregular_models,
                    )
                    eff_cap_day = get_effective_cap_day_after_first_changeover(cap_shift, first_chg_h)
                    cand_eval.append({
                        "line": line,
                        "cap_day": raw_cap_day,
                        "cap_shift": cap_shift,
                        "first_changeover_h": first_chg_h,
                        "effective_cap_day": eff_cap_day,
                    })

                cand_eval = pd.DataFrame(cand_eval)
                positive = cand_eval[cand_eval["effective_cap_day"] > 1e-9].copy()
                if not positive.empty:
                    best = positive.sort_values(
                        ["effective_cap_day", "first_changeover_h", "cap_day"],
                        ascending=[False, True, False]
                    ).iloc[0]
                else:
                    best = cand_eval.sort_values(
                        ["first_changeover_h", "cap_day"],
                        ascending=[True, False]
                    ).iloc[0]
                    warn_rows.append({
                        "date": d.date(),
                        "process": "CUT",
                        "model": model,
                        "part": part,
                        "line": best["line"],
                        "warning_type": "cut_fallback_only_long_changeover_available",
                        "msg": f"无有效正产能候选线，只能回退到首段换线 {best['first_changeover_h']:.2f}h 的线 {best['line']}",
                    })

                cut_assign = pd.concat(
                    [
                        cut_assign,
                        pd.DataFrame([{
                            "date": d.date(),
                            "process": "CUT",
                            "MODEL_NO": model,
                            "PART_NO": part,
                            "key": key,
                            "line": best["line"],
                            "cap_day": best["cap_day"],
                            "carryover": 0,
                            "planned_today_qty": 0.0,
                        }]),
                    ],
                    ignore_index=True,
                )

        if not cut_assign.empty:
            cut_assign = cut_assign.drop_duplicates(subset=["date", "process", "MODEL_NO", "PART_NO", "key", "line"])
            cut_assign = cut_assign.sort_values(["MODEL_NO", "PART_NO", "line"]).reset_index(drop=True)

        for _, r in cut_assign.iterrows():
            model = r["MODEL_NO"]
            part = r["PART_NO"]
            key = r["key"]
            line = r["line"]

            cap_info = cut_cap[(cut_cap["key"] == key) & (cut_cap["line"] == line)]
            if cap_info.empty:
                continue
            cap_shift = float(cap_info["cap_shift"].max())

            prev = line_state["CUT"].get(line, {"model": None, "part": None, "key": None})
            chg_h = get_changeover_h(
                "CUT",
                prev.get("model"),
                prev.get("part"),
                model,
                part,
                changeover,
                line=line,
                hist_recipe_df=hist_recipe,
                irregular_models=irregular_models,
                ratio_is_100=None,
            )

            need_change = chg_h > 0
            chg_shift = "none"

            if need_change:
                if cut_day_chg_used < cut_limits["day"]:
                    chg_shift = "day"
                    cut_day_chg_used += 1
                elif cut_night_chg_used < cut_limits["night"]:
                    chg_shift = "night"
                    cut_night_chg_used += 1
                else:
                    warn_rows.append({
                        "date": d.date(),
                        "process": "CUT",
                        "model": model,
                        "part": part,
                        "line": line,
                        "warning_type": "cut_changeover_limit",
                        "msg": "白班/夜班换线额度都已用完，本线今日不切换",
                    })
                    free_time_cut[line] = 24.0
                    continue

                chg_rows.append({
                    "date": d.date(),
                    "process": "CUT",
                    "line": line,
                    "from_model": prev.get("model"),
                    "from_part": prev.get("part"),
                    "to_model": model,
                    "to_part": part,
                    "changeover_h": chg_h,
                    "shift": chg_shift,
                })

            segments = get_cut_segments(cap_info)
            ratio_need = max(0.0, day_cut_demand_map[key] - day_cut_actual_by_key[key])
            used_h = 0.0
            total_actual = 0.0
            seg_day_qty_total = 0.0
            seg_night_qty_total = 0.0
            ratio_chg_h = get_cut_ratio_changeover_h(model, line, changeover, irregular_models)
            internal_ratio_change_h = 0.0

            for i, seg in enumerate(segments):
                seg_cap_shift = float(seg.get("cap_shift", cap_shift))
                seg_ratio_frac = float(seg.get("ratio_frac", 0.0))
                seg_ratio_pct = float(seg.get("ratio_pct", seg_ratio_frac * 100.0))
                seg_target = ratio_need * seg_ratio_frac
                seg_day_cap, seg_night_cap = plan_shift_qty_with_changeover_shift(
                    seg_cap_shift, chg_h if (i == 0 and need_change) else 0.0, chg_shift if i == 0 else "none"
                )
                if i < len(segments) - 1 and ratio_chg_h > 0:
                    # 中间要留出一次 WITH_RATIO 的时间，优先压缩夜班产能
                    reducible = min(seg_night_cap, seg_cap_shift * (ratio_chg_h / 12.0)) if seg_cap_shift else 0.0
                    seg_night_cap -= reducible
                    left_h = max(0.0, ratio_chg_h - 12.0 * (reducible / seg_cap_shift if seg_cap_shift else 0.0))
                    if left_h > 1e-9 and seg_cap_shift:
                        seg_day_cap = max(0.0, seg_day_cap - seg_cap_shift * (left_h / 12.0))
                    internal_ratio_change_h += ratio_chg_h

                seg_day_qty = min(seg_day_cap, seg_target)
                seg_remain = max(0.0, seg_target - seg_day_qty)
                seg_night_qty = min(seg_night_cap, seg_remain)
                seg_total = seg_day_qty + seg_night_qty

                total_actual += seg_total
                seg_day_qty_total += seg_day_qty
                seg_night_qty_total += seg_night_qty
                used_h += 12.0 * (seg_day_qty / seg_cap_shift if seg_cap_shift else 0.0)
                used_h += 12.0 * (seg_night_qty / seg_cap_shift if seg_cap_shift else 0.0)

                detail_rows.append({
                    "date": d.date(),
                    "process": "CUT",
                    "detail_type": "segment",
                    "event_seq": len(detail_rows) + 1,
                    "MODEL_NO": model,
                    "PART_NO": part,
                    "key": key,
                    "line": line,
                    "segment_no": i + 1,
                    "segment_label": f"第{i + 1}段",
                    "ratio_pct": seg_ratio_pct,
                    "ratio_frac": seg_ratio_frac,
                    "from_model": prev.get("model") if i == 0 else model,
                    "from_part": prev.get("part") if i == 0 else part,
                    "to_model": model,
                    "to_part": part,
                    "shift": chg_shift if (i == 0 and need_change) else "none",
                    "changeover_h": chg_h if (i == 0 and need_change) else 0.0,
                    "day_qty": seg_day_qty,
                    "night_qty": seg_night_qty,
                    "total_qty": seg_total,
                    "cap_shift": seg_cap_shift,
                    "target_qty": seg_target,
                })

                if i == 0 and need_change:
                    used_h += chg_h
                if i < len(segments) - 1:
                    used_h += ratio_chg_h
                    chg_rows.append({
                        "date": d.date(),
                        "process": "CUT",
                        "line": line,
                        "from_model": model,
                        "from_part": part,
                        "to_model": model,
                        "to_part": part,
                        "changeover_h": ratio_chg_h,
                        "shift": "intra",
                    })
                    detail_rows.append({
                        "date": d.date(),
                        "process": "CUT",
                        "detail_type": "intra_changeover",
                        "event_seq": len(detail_rows) + 1,
                        "MODEL_NO": model,
                        "PART_NO": part,
                        "key": key,
                        "line": line,
                        "segment_no": i + 1,
                        "segment_label": f"第{i + 1}段后 WITH_RATIO",
                        "ratio_pct": None,
                        "ratio_frac": None,
                        "from_model": model,
                        "from_part": part,
                        "to_model": model,
                        "to_part": part,
                        "shift": "intra",
                        "changeover_h": ratio_chg_h,
                        "day_qty": 0.0,
                        "night_qty": 0.0,
                        "total_qty": 0.0,
                        "cap_shift": None,
                        "target_qty": target_qty,
                    })

            day_cut_actual_by_key[key] += total_actual
            free_time_cut[line] = max(0.0, 24.0 - used_h)

            if total_actual > 0:
                op_rows.append({
                    "date": d.date(),
                    "process": "CUT",
                    "MODEL_NO": model,
                    "PART_NO": part,
                    "key": key,
                    "line": line,
                    "day_qty": seg_day_qty_total,
                    "night_qty": seg_night_qty_total,
                    "total_qty": total_actual,
                    "cap_shift": cap_shift,
                    "changeover_h": (chg_h if need_change else 0.0) + internal_ratio_change_h,
                })

            if need_change or total_actual > 0:
                line_state["CUT"][line] = {"model": model, "part": part, "key": key}
            if total_actual > 1e-9:
                hist_recipe = append_hist_recipe_if_new(hist_recipe, "CUT", line, model)

        # CUT backlog 输出 + CUT 完工进入 AOI WIP池
        new_cut_backlog = defaultdict(float)
        cut_today_by_key = defaultdict(float)

        for key, total_need in day_cut_demand_map.items():
            actual = day_cut_actual_by_key.get(key, 0.0)
            unmet = max(0.0, total_need - actual)
            model, part = day_prod_map[key]
            cut_today_by_key[(model, part)] += actual

            orig_demand = float(demand[(demand["date"] == d) & (demand["key"] == key)]["qty"].sum())
            carry_in = float(cut_backlog.get(key, 0.0))
            carry_out = 0.0

            if unmet > 1e-9:
                carry_out = unmet
                new_cut_backlog[key] += unmet
                next_prod = get_next_production_date(demand, d)
                if next_prod is not None:
                    warn_rows.append({
                        "date": d.date(),
                        "process": "CUT",
                        "model": model,
                        "part": part,
                        "line": None,
                        "warning_type": "cut_unmet_carry_to_next_production_day",
                        "msg": f"CUT 当天未完成，缺口 {unmet:.2f}，已结转到下一生产日 {next_prod.date()}",
                    })
                else:
                    warn_rows.append({
                        "date": d.date(),
                        "process": "CUT",
                        "model": model,
                        "part": part,
                        "line": None,
                        "warning_type": "cut_unmet_carry_beyond_horizon",
                        "msg": f"CUT 当天未完成，缺口 {unmet:.2f}，已保留为结转，但当前排产窗口内无后续生产日",
                    })

            backlog_rows.append({
                "date": d.date(),
                "process": "CUT",
                "MODEL_NO": model,
                "PART_NO": part,
                "item": key,
                "orig_demand": orig_demand,
                "carry_in": carry_in,
                "actual_qty": actual,
                "unmet": unmet,
                "carry_out": carry_out,
            })

        cut_backlog = new_cut_backlog

        # AOI 期初WIP（仅来自前一生产日及更早的结转）
        aoi_prev_carry_pool = dict(aoi_wip_pool)

        # CUT 当天完工流入 AOI
        same_day_cut_in_map = defaultdict(float)
        for key_mp, qty in cut_today_by_key.items():
            same_day_cut_in_map[key_mp] += qty
            aoi_wip_pool[key_mp] += qty

        # =========================
        # AOI 班次排产（基于真实WIP池）
        # =========================
        aoi_limits = daily_changeover_limits("AOI")
        aoi_day_chg_used = 0
        free_time_aoi = defaultdict(float)
        day_aoi_actual_by_key = defaultdict(float)
        day_aoi_plan_by_key = defaultdict(float)
        # AOI 当天可用WIP = 期初结转 + 当天CUT流入
        initial_aoi_wip_pool = dict(aoi_wip_pool)

        # AOI 的“原始订单需求”仅用于展示与分part优先级（按 model/part 保留）
        raw_aoi_today = (
            demand[demand["date"] == d]
            .groupby(["MODEL_NO", "PART_NO"], as_index=False)["qty"]
            .sum()
        )
        raw_aoi_need_map = defaultdict(float)
        for _, r in raw_aoi_today.iterrows():
            raw_aoi_need_map[(r["MODEL_NO"], r["PART_NO"])] += float(r["qty"])

        # 关键修复：
        # AOI 同一 model 多 part 时，必须基于“具体part的需求/结转”重新分配到 line。
        # 优先让历史相同 part 的线继续跑原 part，并尽量保证每个有需求/有结转的 part 至少拿到 1 条线。
        if not aoi_assign.empty:
            aoi_assign = aoi_assign.drop_duplicates(subset=["date", "process", "MODEL_NO", "PART_NO", "line"])
            aoi_assign = aoi_assign.sort_values(["MODEL_NO", "PART_NO", "line"]).reset_index(drop=True)

            reassigned_rows = []
            for model, grp in aoi_assign.groupby("MODEL_NO", dropna=False):
                chosen_lines = []
                for _, rr in grp.iterrows():
                    chosen_lines.append((
                        rr["line"],
                        float(rr.get("cap_day", 0.0) if pd.notna(rr.get("cap_day", 0.0)) else 0.0),
                        float(rr.get("effective_cap_day", rr.get("cap_day", 0.0)) if pd.notna(rr.get("effective_cap_day", rr.get("cap_day", 0.0))) else 0.0),
                        float(rr.get("first_changeover_h", 0.0) if pd.notna(rr.get("first_changeover_h", 0.0)) else 0.0),
                        float(rr.get("effective_cap_day", rr.get("cap_day", 0.0)) if pd.notna(rr.get("effective_cap_day", rr.get("cap_day", 0.0))) else 0.0),
                        int(rr.get("carryover", 0) if pd.notna(rr.get("carryover", 0)) else 0),
                    ))

                part_need_map = {}
                parts_for_model = {
                    p for (m, p), q in raw_aoi_need_map.items()
                    if m == model and q > 1e-9
                } | {
                    p for (m, p), q in aoi_prev_carry_pool.items()
                    if m == model and q > 1e-9
                } | {
                    p for (m, p), q in same_day_cut_in_map.items()
                    if m == model and q > 1e-9
                }

                for part in parts_for_model:
                    # 分part时，优先级看“期初结转 + 当天原始需求 + 当天CUT流入”
                    part_need_map[part] = (
                        float(aoi_prev_carry_pool.get((model, part), 0.0))
                        + float(raw_aoi_need_map.get((model, part), 0.0))
                        + float(same_day_cut_in_map.get((model, part), 0.0))
                    )

                line_part_map = assign_aoi_parts_to_lines(model, chosen_lines, part_need_map, line_state)

                remaining_plan = {str(p).strip(): float(q) for p, q in part_need_map.items()}
                for _, rr in grp.sort_values(["carryover", "line"], ascending=[False, True]).iterrows():
                    line = rr["line"]
                    part = line_part_map.get(line, rr.get("PART_NO"))
                    eff_cap = float(rr.get("effective_cap_day", rr.get("cap_day", 0.0)) if pd.notna(rr.get("effective_cap_day", rr.get("cap_day", 0.0))) else 0.0)
                    planned_today_qty = min(eff_cap, max(0.0, remaining_plan.get(part, 0.0)))
                    remaining_plan[part] = max(0.0, remaining_plan.get(part, 0.0) - planned_today_qty)

                    rr2 = rr.copy()
                    rr2["PART_NO"] = part
                    rr2["planned_today_qty"] = planned_today_qty
                    reassigned_rows.append(rr2)

            if reassigned_rows:
                aoi_assign = pd.DataFrame(reassigned_rows).reset_index(drop=True)
                aoi_assign = aoi_assign.sort_values(["MODEL_NO", "PART_NO", "line"]).reset_index(drop=True)

        for _, r in aoi_assign.iterrows():
            model = r["MODEL_NO"]
            line = r["line"]

            cap_info = aoi_cap[(aoi_cap["MODEL_NO"] == model) & (aoi_cap["line"] == line)]
            if cap_info.empty:
                continue
            cap_shift = float(cap_info["cap_shift"].max())

            prev = line_state["AOI"].get(line, {"model": None, "part": None})
            aoi_part = r.get("PART_NO") if isinstance(r, pd.Series) else r["PART_NO"]
            if pd.isna(aoi_part) or aoi_part is None or str(aoi_part).strip() == "":
                aoi_part = day_model_part_map.get(model, model_part_ref.get(model))
            chg_h = get_changeover_h(
                "AOI",
                prev.get("model"),
                prev.get("part"),
                model,
                aoi_part,
                changeover,
                line=line,
                hist_recipe_df=hist_recipe,
                irregular_models=irregular_models,
                ratio_is_100=None,
            )

            need_change = chg_h > 0
            if need_change:
                if aoi_day_chg_used >= aoi_limits["day"]:
                    warn_rows.append({
                        "date": d.date(),
                        "process": "AOI",
                        "model": model,
                        "part": day_model_part_map.get(model, model_part_ref.get(model)),
                        "line": line,
                        "warning_type": "aoi_changeover_limit",
                        "msg": "换线超限，优先守住换线数量限制，本线今日不切换",
                    })
                    free_time_aoi[line] = 24.0
                    continue

                aoi_day_chg_used += 1
                chg_rows.append({
                    "date": d.date(),
                    "process": "AOI",
                    "line": line,
                    "from_model": prev.get("model"),
                    "from_part": prev.get("part"),
                    "to_model": model,
                    "to_part": aoi_part,
                    "changeover_h": chg_h,
                    "shift": "day",
                })

            day_qty_cap, night_qty_cap = plan_shift_qty_with_changeover_shift(
                cap_shift, chg_h if need_change else 0.0, "day" if need_change else "none"
            )

            # AOI 只从真实 WIP池取量（按 model/part）
            mp_key = (model, aoi_part)
            remain_need_before = max(0.0, initial_aoi_wip_pool.get(mp_key, 0.0) - day_aoi_plan_by_key[mp_key])
            target_qty = min(day_qty_cap + night_qty_cap, remain_need_before)
            day_aoi_plan_by_key[mp_key] += target_qty

            remain_need = max(0.0, aoi_wip_pool.get(mp_key, 0.0) - day_aoi_actual_by_key[mp_key])

            day_qty = min(day_qty_cap, remain_need)
            remain_need -= day_qty
            night_qty = min(night_qty_cap, remain_need)

            total_actual = day_qty + night_qty
            day_aoi_actual_by_key[mp_key] += total_actual

            used_h = (chg_h if need_change else 0.0) + 12.0 * (day_qty / cap_shift if cap_shift else 0.0) + 12.0 * (night_qty / cap_shift if cap_shift else 0.0)
            free_time_aoi[line] = max(0.0, 24.0 - used_h)

            if total_actual > 0:
                op_rows.append({
                    "date": d.date(),
                    "process": "AOI",
                    "MODEL_NO": model,
                    "PART_NO": aoi_part,
                    "key": None,
                    "line": line,
                    "day_qty": day_qty,
                    "night_qty": night_qty,
                    "total_qty": total_actual,
                    "cap_shift": cap_shift,
                    "changeover_h": chg_h if need_change else 0.0,
                })
                detail_rows.append({
                    "date": d.date(),
                    "process": "AOI",
                    "detail_type": "production",
                    "event_seq": len(detail_rows) + 1,
                    "MODEL_NO": model,
                    "PART_NO": aoi_part,
                    "key": None,
                    "line": line,
                    "segment_no": 1,
                    "segment_label": "生产",
                    "ratio_pct": None,
                    "ratio_frac": None,
                    "from_model": prev.get("model"),
                    "from_part": prev.get("part"),
                    "to_model": model,
                    "to_part": aoi_part,
                    "shift": "day" if need_change else "none",
                    "changeover_h": chg_h if need_change else 0.0,
                    "day_qty": day_qty,
                    "night_qty": night_qty,
                    "total_qty": total_actual,
                    "cap_shift": cap_shift,
                    "orig_demand": raw_aoi_need_map.get((model, aoi_part), 0.0),
                    "carry_in": aoi_prev_carry_pool.get((model, aoi_part), 0.0),
                    "same_day_cut_in": same_day_cut_in_map.get((model, aoi_part), 0.0),
                    "available_wip": initial_aoi_wip_pool.get((model, aoi_part), 0.0),
                    "target_qty": target_qty,
                })

            if need_change or total_actual > 0:
                line_state["AOI"][line] = {"model": model, "part": aoi_part, "key": None}
            if total_actual > 1e-9:
                hist_recipe = append_hist_recipe_if_new(hist_recipe, "AOI", line, model)

        # AOI 消耗 WIP池 + backlog输出（按 model/part）
        for mp_key, actual in day_aoi_actual_by_key.items():
            aoi_wip_pool[mp_key] = max(0.0, aoi_wip_pool.get(mp_key, 0.0) - actual)

        all_mp_keys = set(raw_aoi_need_map.keys()) | set(aoi_wip_pool.keys()) | set(day_aoi_actual_by_key.keys())

        for mp_key in sorted(all_mp_keys):
            model, aoi_part = mp_key
            actual = day_aoi_actual_by_key.get(mp_key, 0.0)
            orig_demand = raw_aoi_need_map.get(mp_key, 0.0)
            carry_in = aoi_prev_carry_pool.get(mp_key, 0.0)
            same_day_cut_in = same_day_cut_in_map.get(mp_key, 0.0)
            available_wip = initial_aoi_wip_pool.get(mp_key, 0.0)
            unmet = aoi_wip_pool.get(mp_key, 0.0)  # 剩余WIP
            carry_out = unmet

            if unmet > 1e-9:
                warn_rows.append({
                    "date": d.date(),
                    "process": "AOI",
                    "model": model,
                    "part": aoi_part,
                    "line": None,
                    "warning_type": "aoi_unmet_auto_roll",
                    "msg": f"AOI 当天未完成，WIP剩余 {unmet:.2f}，已自动滚到次日",
                })

            if actual < orig_demand - 1e-9:
                warn_rows.append({
                    "date": d.date(),
                    "process": "AOI",
                    "model": model,
                    "part": aoi_part,
                    "line": None,
                    "warning_type": "aoi_raw_demand_not_met",
                    "msg": f"原始当日需求 {orig_demand:.2f}，实际仅排 {actual:.2f}",
                })

            backlog_rows.append({
                "date": d.date(),
                "process": "AOI",
                "MODEL_NO": model,
                "PART_NO": aoi_part,
                "item": model,
                "orig_demand": orig_demand,
                "carry_in": carry_in,
                "same_day_cut_in": same_day_cut_in,
                "available_wip": available_wip,
                "actual_qty": actual,
                "unmet": unmet,
                "carry_out": carry_out,
            })

        # ------------ 日汇总 ------------
        ops_df_day = pd.DataFrame(op_rows)
        ops_df_day = ops_df_day[ops_df_day["date"] == d.date()].copy()

        if not ops_df_day.empty:
            cut_sum = (
                ops_df_day[ops_df_day["process"] == "CUT"]
                .groupby(["process", "MODEL_NO", "PART_NO"], as_index=False)[["day_qty", "night_qty", "total_qty"]]
                .sum()
            )
            aoi_sum = (
                ops_df_day[ops_df_day["process"] == "AOI"]
                .groupby(["process", "MODEL_NO", "PART_NO"], as_index=False)[["day_qty", "night_qty", "total_qty"]]
                .sum()
            )

            if not cut_sum.empty:
                cut_sum["date"] = d.date()
                daily_rows.extend(
                    cut_sum[["date", "process", "MODEL_NO", "PART_NO", "day_qty", "night_qty", "total_qty"]].to_dict("records")
                )

            if not aoi_sum.empty:
                aoi_sum["date"] = d.date()
                daily_rows.extend(
                    aoi_sum[["date", "process", "MODEL_NO", "PART_NO", "day_qty", "night_qty", "total_qty"]].to_dict("records")
                )

        # ------------ 预换线 ------------
        if d >= pd.Timestamp("2026-02-04"):
            next_date = get_next_production_date(demand, d)
            if next_date is not None:
                cut_remaining_changeovers_today = max(0, daily_changeover_limits("CUT")["day"] - cut_day_chg_used)
                cut_sugs = try_prechangeover(
                    current_date=d,
                    next_date=next_date,
                    process="CUT",
                    line_state=line_state,
                    free_time_today_by_line=free_time_cut,
                    remaining_changeovers_today=cut_remaining_changeovers_today,
                    demand=demand,
                    cut_cap=cut_cap,
                    aoi_cap=aoi_cap,
                    changeover=changeover,
                    hist_recipe=hist_recipe,
                    irregular_models=irregular_models,
                )
                prechg_rows.extend(cut_sugs)

                aoi_remaining_changeovers_today = max(0, daily_changeover_limits("AOI")["day"] - aoi_day_chg_used)
                aoi_sugs = try_prechangeover(
                    current_date=d,
                    next_date=next_date,
                    process="AOI",
                    line_state=line_state,
                    free_time_today_by_line=free_time_aoi,
                    remaining_changeovers_today=aoi_remaining_changeovers_today,
                    demand=demand,
                    cut_cap=cut_cap,
                    aoi_cap=aoi_cap,
                    changeover=changeover,
                    hist_recipe=hist_recipe,
                    irregular_models=irregular_models,
                )
                prechg_rows.extend(aoi_sugs)

    daily_df = pd.DataFrame(daily_rows)
    ops_df = pd.DataFrame(op_rows)
    detail_df = pd.DataFrame(detail_rows)
    changes_df = pd.DataFrame(chg_rows)
    warnings_df = pd.DataFrame(warn_rows)
    backlog_df = pd.DataFrame(backlog_rows)
    prechg_df = pd.DataFrame(prechg_rows)
    master_df = build_master_schedule_table(ops_df, backlog_df, warnings_df)
    all_changes_df = build_all_changeovers_table(changes_df, prechg_df)
    master_detail_df = build_master_schedule_detail_table(detail_df, backlog_df, warnings_df)

    return daily_df, changes_df, ops_df, detail_df, warnings_df, backlog_df, prechg_df, master_df, master_detail_df, all_changes_df



# =========================================================
# main
# =========================================================

if __name__ == "__main__":
    from pathlib import Path

    base_dir = Path(__file__).resolve().parent

    data = load_data(
        demand_path=str(base_dir / "demand_202603.csv"),
        hist_path=str(base_dir / "hist_260313.csv"),
        aoi_cap_path=str(base_dir / "aoi_cap.csv"),
        cut_cap_path=str(base_dir / "cut_cap.csv"),
        changeover_path=str(base_dir / "changeover.csv"),
        hist_recipe_path=str(base_dir / "hist_recipe.csv"),
        irregular_path=str(base_dir / "irregular.csv"),
        year=2026,
    )

    step0_summary, step0_warn = step0_checks(data)

    (
        daily_df,
        changes_df,
        ops_df,
        detail_df,
        warnings_df,
        backlog_df,
        prechg_df,
        master_df,
        master_detail_df,
        all_changes_df,
    ) = run_step2_with_backlog(
        data,
        start_date="2026-03-14",
        end_date="2026-03-31",
    )

    logic_check = pd.DataFrame([
        {
            "item": "Step1分线（CUT/AOI）",
            "current_logic": "按首日有效产能（cap_day扣首段换线）+ 连续需求保留线 + segment total 分线",
            "uses_changeover": "是",
            "only_first_changeover": "是",
            "conclusion": "当前选线已考虑首段换线时间，但不会把CUT内部 WITH_RATIO 算入选线阶段",
        },
        {
            "item": "pre-changeover（提前换线建议）",
            "current_logic": "只调用 get_changeover_h 判断能否提前切到次日目标",
            "uses_changeover": "是",
            "only_first_changeover": "是",
            "conclusion": "这里只看首段换线时间，不会把 CUT 内部 WITH_RATIO 算进去",
        },
        {
            "item": "CUT当天实际产能扣减",
            "current_logic": "先扣首段换线 chg_h，再对多 ratio 产品在段间扣 WITH_RATIO",
            "uses_changeover": "是",
            "only_first_changeover": "否",
            "conclusion": "这一步不是选 line，而是已选定 line 后计算当天还能生产多少",
        },
        {
            "item": "hist空机台状态",
            "current_logic": "hist_260313 中若某 line 的 MODEL/PART 为空，则显式标记为 IDLE",
            "uses_changeover": "是",
            "only_first_changeover": "是",
            "conclusion": "IDLE line 不参与保留，可按目标产品直接切换使用；选线时仍只计算切入目标产品的首段换线",
        },
    ])

    summary_df = pd.DataFrame([{
        "demand_start": str(data["demand"]["date"].min().date()),
        "demand_end": str(data["demand"]["date"].max().date()),
        "run_start": "2026-03-14",
        "run_end": "2026-03-31",
        "cut_detail_rows": len(detail_df[detail_df["process"] == "CUT"]),
        "aoi_detail_rows": len(detail_df[detail_df["process"] == "AOI"]),
        "operations_rows": len(ops_df),
        "warnings_rows": len(warnings_df),
        "backlog_rows": len(backlog_df),
        "prechg_rows": len(prechg_df),
    }])

    sheets = {
        "summary": summary_df,
        "logic_check": logic_check,
        "master_schedule_detail": master_detail_df,
        "cut_detail": detail_df[detail_df["process"] == "CUT"].reset_index(drop=True),
        "ops_detail_all": detail_df,
        "master_schedule": master_df,
        "operations": ops_df,
        "all_changeovers": all_changes_df,
        "warnings": warnings_df,
        "backlog": backlog_df,
        "daily_summary": daily_df,
        "prechangeovers": prechg_df,
        "step0_summary": step0_summary,
        "step0_warnings": step0_warn,
    }

    output_xlsx = base_dir / "march_schedule_results_detailed.xlsx"
    export_results_to_excel(str(output_xlsx), sheets)

    print(f"Saved workbook: {output_xlsx}")
